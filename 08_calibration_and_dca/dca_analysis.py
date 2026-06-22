"""Decision Curve Analysis (DCA) for DR referral endpoints.

Implements net benefit curves for binary endpoints derived from DR grades:
- endpoint y>=k (e.g., k=2 for RDR, k=3 for STDR)

Inputs:
- A CSV containing y_true (0-4) and class probabilities p0..p4.
  Supported formats:
  - preds_image_*_test*.csv from eval_grade_ce_posthoc.py (true_grade, p0..p4)
  - ensemble_preds_*.csv from external_ensemble_eval.py (y_true, p0..p4)

Outputs:
- curves CSV and publication-ready PNG/PDF
- optional Excel export (.xlsx) with curves + summary table
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import matplotlib
import matplotlib.pyplot as plt
import numpy as np


def _try_import_openpyxl():
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, Alignment  # type: ignore
        return openpyxl, Font, Alignment
    except Exception as e:
        raise RuntimeError(
            "Excel export requires 'openpyxl'. Install it in your environment, e.g. `pip install openpyxl`. "
            f"Original error: {e}"
        )


def read_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def get_y_true(row: Dict[str, str]) -> int:
    for k in ("true_grade", "y_true"):
        if k in row and row[k] not in ("", None):
            return int(float(row[k]))
    raise KeyError("Missing y_true/true_grade column")


def get_probs(row: Dict[str, str]) -> np.ndarray:
    ps = []
    for i in range(5):
        v = row.get(f"p{i}", "")
        if v in ("", None):
            raise KeyError("Missing p0..p4 columns")
        ps.append(float(v))
    p = np.array(ps, dtype=np.float64)
    p = np.clip(p, 1e-12, 1.0)
    p = p / p.sum()
    return p


def net_benefit(y_true01: np.ndarray, p_hat: np.ndarray, pt: float) -> float:
    """NB(pt) = TP/N - FP/N * pt/(1-pt)."""
    y = np.asarray(y_true01).astype(int)
    p = np.asarray(p_hat).astype(float)
    m = np.isfinite(p) & ((y == 0) | (y == 1))
    y = y[m]
    p = p[m]
    N = int(y.size)
    if N == 0:
        return float("nan")
    # Handle endpoints for nicer plots:
    # - pt<=0: everyone is positive -> NB = event_rate
    # - pt>=1: nobody is positive -> NB = 0
    if pt <= 0.0:
        return float((y == 1).sum() / N)
    if pt >= 1.0:
        return 0.0
    pred = (p >= pt).astype(int)
    TP = int(((pred == 1) & (y == 1)).sum())
    FP = int(((pred == 1) & (y == 0)).sum())
    w = pt / (1.0 - pt)
    return float(TP / N - FP / N * w)


def treat_all_nb(y_true01: np.ndarray, pt: float) -> float:
    y = np.asarray(y_true01).astype(int)
    m = (y == 0) | (y == 1)
    y = y[m]
    N = int(y.size)
    if N == 0:
        return float("nan")
    if pt <= 0.0:
        return float((y == 1).sum() / N)
    if pt >= 1.0:
        return float("nan")
    TP = int((y == 1).sum())
    FP = int((y == 0).sum())
    w = pt / (1.0 - pt)
    return float(TP / N - FP / N * w)


def treat_none_nb(_: np.ndarray, __: float) -> float:
    return 0.0


def set_style() -> None:
    try:
        matplotlib.use("Agg", force=True)
    except Exception:
        pass
    plt.rcParams.update(
        {
            "figure.dpi": 150,
            "savefig.dpi": 300,
            "font.family": "serif",
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "axes.titlesize": 12,
            "axes.labelsize": 11,
            "xtick.labelsize": 10,
            "ytick.labelsize": 10,
            "legend.fontsize": 9,
            "axes.grid": True,
            "grid.alpha": 0.25,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "pdf.fonttype": 42,  # TrueType fonts (editable in Illustrator)
            "ps.fonttype": 42,   # TrueType fonts for PS/EPS
        }
    )


def export_excel(out_xlsx: Path, curves: List[Dict[str, Any]], summary: List[Dict[str, Any]]) -> None:
    openpyxl, Font, Alignment = _try_import_openpyxl()
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "dca_curves"
    headers = ["endpoint", "pt", "nb_model", "nb_treat_all", "nb_treat_none"]
    ws0.append(headers)
    for r in curves:
        ws0.append([r["endpoint"], float(r["pt"]), float(r["nb_model"]), float(r["nb_treat_all"]), float(r["nb_treat_none"])])

    ws1 = wb.create_sheet("dca_summary")
    ws1.append(["endpoint", "pt_range", "max_nb_pt", "max_nb", "event_rate"])
    for r in summary:
        ws1.append([r["endpoint"], r["pt_range"], float(r["max_nb_pt"]), float(r["max_nb"]), float(r["event_rate"])])

    header_font = Font(bold=True)
    for ws in (ws0, ws1):
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = 0
            for c in col_cells[:200]:
                v = c.value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 42)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_xlsx))


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--in_csv", type=str, required=True)
    p.add_argument("--out_dir", type=str, default="checkpoints/dca")
    p.add_argument("--tag", type=str, default="internal")
    p.add_argument("--endpoints", type=str, default="2,3", help="Comma-separated k for y>=k endpoints")
    p.add_argument("--pt_min", type=float, default=0.0)
    p.add_argument("--pt_max", type=float, default=1.0)
    p.add_argument("--pt_step", type=float, default=0.01)
    p.add_argument("--excel", type=int, default=1, choices=[0, 1])
    p.add_argument(
        "--ytrim_q",
        type=float,
        default=0.01,
        help="Lower-quantile clipping for y-axis (match LR,SVM notebook style).",
    )
    p.add_argument("--dpi", type=int, default=600, help="Figure DPI for PNG/PDF export")
    args = p.parse_args()

    in_csv = Path(args.in_csv).resolve()
    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    tag = str(args.tag).strip() or "dca"
    ks = [int(x) for x in str(args.endpoints).split(",") if x.strip()]
    pt_min = float(args.pt_min)
    pt_max = float(args.pt_max)
    pt_step = float(args.pt_step)
    ytrim_q = float(args.ytrim_q)
    dpi = int(args.dpi)
    if not (0.0 <= pt_min < pt_max <= 1.0):
        raise ValueError("Require 0 <= pt_min < pt_max <= 1")
    if pt_step <= 0:
        raise ValueError("pt_step must be > 0")
    if not (0.0 <= ytrim_q < 0.5):
        raise ValueError("Require 0 <= ytrim_q < 0.5")
    if dpi <= 0:
        raise ValueError("dpi must be > 0")

    rows = read_rows(in_csv)
    y = np.array([get_y_true(r) for r in rows], dtype=np.int64)
    probs = np.stack([get_probs(r) for r in rows], axis=0).astype(np.float64)

    set_style()

    eps = 1e-6
    pts_core = np.arange(max(eps, pt_min), min(pt_max, 1.0 - eps) + 1e-12, pt_step, dtype=np.float64)
    pts_list: List[float] = []
    if pt_min <= 0.0:
        pts_list.append(0.0)
    pts_list.extend([float(x) for x in pts_core.tolist()])
    if pt_max >= 1.0:
        pts_list.append(1.0)
    pts = np.asarray(pts_list, dtype=np.float64)
    all_curves: List[Dict[str, Any]] = []
    summary: List[Dict[str, Any]] = []

    for k in ks:
        if k < 1 or k > 4:
            raise ValueError("Endpoint k must be in [1..4]")
        y01 = (y >= k).astype(int)
        p_hat = probs[:, k:].sum(axis=1)
        ev_rate = float(y01.mean())

        nb_model = np.array([net_benefit(y01, p_hat, float(pt)) for pt in pts], dtype=np.float64)
        nb_all = np.array([treat_all_nb(y01, float(pt)) for pt in pts], dtype=np.float64)
        nb_none = np.zeros_like(nb_model)

        for i, pt in enumerate(pts.tolist()):
            all_curves.append({"endpoint": f"y>= {k}", "pt": float(pt), "nb_model": float(nb_model[i]), "nb_treat_all": float(nb_all[i]), "nb_treat_none": 0.0})

        # max NB within range
        j = int(np.nanargmax(nb_model)) if np.isfinite(nb_model).any() else 0
        summary.append({
            "endpoint": f"y>= {k}",
            "pt_range": f"[{pt_min:.2f}, {pt_max:.2f}]",
            "max_nb_pt": float(pts[j]),
            "max_nb": float(nb_model[j]),
            "event_rate": ev_rate,
        })

        # plot (match LR,SVM建模.ipynb DCA style)
        # ========== SCI-STYLE DCA PLOT ==========
        fig = plt.figure(figsize=(8.5, 6.0))
        ax = plt.gca()
        
        # Treat none baseline (dotted gray)
        ax.plot((0, 1), (0, 0), linestyle=':', color='#7F8C8D', linewidth=1.5, 
                label='Treat None', zorder=2)
        
        # Treat all (dashed black) - shared reference line
        ax.plot(pts, nb_all, color='#2C3E50', linestyle='--', linewidth=2.0, 
                label='Treat All', zorder=3)
        
        # Model curve (solid colored)
        ax.plot(pts, nb_model, color='#3498DB', linewidth=2.5, 
                label='Model', zorder=4)

        # Benefit area with refined styling
        base = np.maximum(nb_all, 0.0)
        top = np.maximum(nb_model, base)
        ax.fill_between(pts, top, base, alpha=0.2, color='#2ECC71', 
                       label='Net Benefit', zorder=1)

        # y-limits: FOCUS ON MODEL CURVE, cap extremes from Treat-all
        if nb_model.size == 0:
            y_min, y_max = -0.05, 0.05
        else:
            # Use model curve to determine range (ignore treat-all extremes)
            model_min = float(np.percentile(nb_model[np.isfinite(nb_model)], 2)) if np.isfinite(nb_model).any() else 0
            model_max = float(np.percentile(nb_model[np.isfinite(nb_model)], 98)) if np.isfinite(nb_model).any() else 0
            
            # Expand range slightly for context
            y_range = model_max - model_min
            y_min = max(model_min - 0.15 * y_range, -0.05)  # Don't go too negative
            y_max = min(model_max + 0.20 * y_range, 0.50)   # Cap upper limit
            
            # Ensure y=0 (Treat None) is visible
            y_min = min(y_min, -0.02)
        
        pad = 0.02
        ax.set_ylim(y_min - pad, y_max + pad)

        ax.set_xlabel("Threshold Probability", fontsize=11, fontweight='medium')
        ax.set_ylabel("Net Benefit", fontsize=11, fontweight='medium')
        
        # Endpoint name mapping
        endpoint_names = {2: "RDR (y≥2)", 3: "STDR (y≥3)"}
        ax.set_title(f"Decision Curve Analysis: {endpoint_names.get(k, f'y≥{k}')}\n({tag})", 
                    fontsize=12, fontweight='bold', pad=10)
        ax.set_xlim(0.0, 1.0)
        
        # Refined legend
        ax.legend(loc='upper right', frameon=True, fancybox=True, shadow=False, 
                 fontsize=9, edgecolor='#CCCCCC', facecolor='white')
        
        # Grid and spines
        ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
        ax.set_axisbelow(True)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(1.2)
        ax.spines['bottom'].set_linewidth(1.2)
        
        fig.tight_layout()

        out_pdf = out_dir / f"dca_{tag}_yge{k}.pdf"
        out_png = out_dir / f"dca_{tag}_yge{k}.png"
        plt.savefig(out_pdf, dpi=dpi, bbox_inches="tight")
        plt.savefig(out_png, dpi=dpi, bbox_inches="tight")
        plt.close(fig)

    curves_csv = out_dir / f"dca_curves_{tag}.csv"
    with curves_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["endpoint", "pt", "nb_model", "nb_treat_all", "nb_treat_none"])
        w.writeheader()
        for r in all_curves:
            w.writerow(r)

    out_json = out_dir / f"dca_summary_{tag}.json"
    with out_json.open("w", encoding="utf-8") as f:
        json.dump({"in_csv": str(in_csv), "endpoints": ks, "pt_min": pt_min, "pt_max": pt_max, "pt_step": pt_step, "summary": summary}, f, indent=2, ensure_ascii=False)

    if int(args.excel) == 1:
        out_xlsx = out_dir / f"dca_{tag}.xlsx"
        export_excel(out_xlsx, all_curves, summary)
        print("[OK] wrote:", out_xlsx)

    print("[OK] wrote:", curves_csv)
    print("[OK] wrote:", out_json)
    print("[OK] figures saved in:", out_dir)


if __name__ == "__main__":
    main()
