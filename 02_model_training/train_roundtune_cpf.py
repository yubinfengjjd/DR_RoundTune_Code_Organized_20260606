# -*- coding: utf-8 -*-
"""
Quality-aware RoundTune training script (quality pretraining + image-only DR grading).

What you can do NOW:
  - Train image quality/gradability head:          --task qual
  - Train image-only DR grading head:              --task grade

Key design:
  - One shared image encoder (backbone/encoder) initialized from RoundTune weights.
  - Separate quality and DR grading heads on the shared image encoder.
  - Loss masking: missing labels are skipped for mixed manifests.

Outputs per run:
  <PROJECT>/checkpoints/runs/<task>/
    best_*.pth
    best_encoder.pth
    logs/history.csv, logs/history.jsonl, logs/summary.json
    tables/metrics_summary.csv, tables/metrics_summary.xlsx (if openpyxl installed)
    tb/  (if --tensorboard enabled)

Repository organization note: this script is the training entry point for
quality pretraining and image-level DR grading.

English comments only.
"""
from __future__ import annotations

# ---- Windows GBK console safe output (avoid UnicodeEncodeError) ----
import sys as _sys
try:
    if hasattr(_sys.stdout, 'reconfigure'):
        _sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(_sys.stderr, 'reconfigure'):
        _sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass
# -------------------------------------------------------------------

# ---- Windows DLL conflict workaround: preload Pillow core before timm/torchvision ----
try:
    from PIL import _imaging  # noqa: F401
except Exception:
    pass

import argparse
import csv
import json
import math
import random
import re
import time
import sys
from contextlib import nullcontext
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader, Dataset
from torchvision import transforms as T

# Optional deps
try:
    import cv2  # type: ignore
except Exception:
    cv2 = None

try:
    import albumentations as A  # type: ignore
    from albumentations.pytorch import ToTensorV2  # type: ignore
except Exception:
    A = None
    ToTensorV2 = None

try:
    import timm  # type: ignore
except Exception:
    timm = None

try:
    from sklearn.model_selection import StratifiedShuffleSplit, GroupShuffleSplit  # type: ignore
    from sklearn.metrics import cohen_kappa_score, f1_score  # type: ignore
except Exception:
    StratifiedShuffleSplit = None
    GroupShuffleSplit = None
    cohen_kappa_score = None
    f1_score = None

try:
    from torch.utils.tensorboard import SummaryWriter  # type: ignore
except Exception:
    SummaryWriter = None

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
except Exception:
    openpyxl = None
    Alignment = Font = PatternFill = None



# -----------------------------
# Utilities
# -----------------------------
def seed_everything(seed: int = 42) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def _device_type(device: str) -> str:
    d = str(device).lower()
    return "cuda" if d.startswith("cuda") else "cpu"


def amp_autocast(device: str, enabled: bool):
    if not enabled:
        return nullcontext()
    dev = _device_type(device)
    if hasattr(torch, "amp") and hasattr(torch.amp, "autocast"):
        return torch.amp.autocast(dev)
    return torch.cuda.amp.autocast()


def amp_grad_scaler(device: str, enabled: bool):
    dev = _device_type(device)
    enabled = bool(enabled) and dev == "cuda"
    if hasattr(torch, "amp") and hasattr(torch.amp, "GradScaler"):
        return torch.amp.GradScaler("cuda", enabled=enabled)
    return torch.cuda.amp.GradScaler(enabled=enabled)


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in r.items()})
    if not rows:
        raise ValueError(f"Empty CSV: {path}")
    return rows


def resolve_path(project_root: Path, p: str) -> str:
    pp = Path(p)
    if pp.is_absolute():
        return str(pp)
    return str((project_root / pp).resolve())


def strip_module_prefix(state: Dict[str, torch.Tensor]) -> Dict[str, torch.Tensor]:
    out: Dict[str, torch.Tensor] = {}
    for k, v in state.items():
        if k.startswith("module."):
            out[k[len("module."):]] = v
        else:
            out[k] = v
    return out


def infer_backbone_from_state_dict_keys(keys: List[str], state: Optional[Dict[str, torch.Tensor]] = None) -> Optional[str]:
    def _find_state_key(suffix: str) -> Optional[str]:
        if state is not None:
            if suffix in state:
                return suffix
            # Common wrappers: backbone./encoder./model./module. etc. Use suffix match.
            for k in state.keys():
                if k.endswith(suffix):
                    return k
            return None
        for k in keys:
            if k.endswith(suffix):
                return k
        return None

    kset = set(keys)
    if any("downsample_layers" in k for k in kset) and any("stages" in k for k in kset):
        return "convnext_tiny"
    if any("mask_token" in k for k in kset) and any("decoder_embed" in k for k in kset):
        # MAE-style ViT (encoder + decoder)
        pe = _find_state_key("patch_embed.proj.weight")
        if state is not None and pe is not None:
            embed_dim = state[pe].shape[0]
            return "mae_vit_large_patch16_224" if embed_dim >= 1000 else "mae_vit_base_patch16_224"
        return "mae_vit_base_patch16_224"
    if any("patch_embed" in k for k in kset) and any("relative_position_bias_table" in k for k in kset):
        # Swin-style
        return "swin_tiny_patch4_window7_224"
    if any("patch_embed.proj" in k for k in kset) and any(re.search(r"\bblocks\.\d+\.", k) for k in kset):
        # ViT-style; pick variant by embed dim if available
        pe = _find_state_key("patch_embed.proj.weight")
        if state is not None and pe is not None:
            embed_dim = state[pe].shape[0]
            if embed_dim >= 1000:
                return "vit_large_patch16_224"
            if embed_dim >= 700:
                return "vit_base_patch16_224"
        return "vit_base_patch16_224"
    return None


def load_checkpoint_state(ckpt_path: Path) -> Dict[str, torch.Tensor]:
    """Load a checkpoint in a way that is compatible with PyTorch >=2.6 security defaults.

    PyTorch 2.6+ flips torch.load(weights_only=...) default to True, which may fail for
    checkpoints that contain harmless metadata objects (e.g. argparse.Namespace).
    We try the safe path first (weights_only=True with an allowlist), and fall back to
    weights_only=False only if needed and the checkpoint is trusted.
    """
    import pickle
    import argparse

    # 1) Try the safe default: weights_only=True (PyTorch >=2.6 default)
    try:
        obj = torch.load(str(ckpt_path), map_location="cpu", weights_only=True)
    except TypeError:
        # Older PyTorch without weights_only argument
        obj = torch.load(str(ckpt_path), map_location="cpu")
    except pickle.UnpicklingError:
        # 2) Allowlist common harmless metadata types seen in training checkpoints
        try:
            from torch.serialization import safe_globals  # PyTorch >=2.4
            with safe_globals([argparse.Namespace]):
                obj = torch.load(str(ckpt_path), map_location="cpu", weights_only=True)
        except Exception:
            # 3) Last resort: load the full pickle (ONLY if you trust the checkpoint source)
            obj = torch.load(str(ckpt_path), map_location="cpu", weights_only=False)

    if isinstance(obj, dict):
        if "state_dict" in obj and isinstance(obj["state_dict"], dict):
            state = obj["state_dict"]
        elif "model" in obj and isinstance(obj["model"], dict):
            state = obj["model"]
        else:
            state = {k: v for k, v in obj.items() if torch.is_tensor(v)}
    else:
        raise ValueError(f"Unsupported checkpoint format: {ckpt_path}")
    return strip_module_prefix(state)



def load_roundtune_encoder_weights(ckpt_path: Path) -> Tuple[Dict[str, torch.Tensor], Optional[str]]:
    state = load_checkpoint_state(ckpt_path)
    inferred = infer_backbone_from_state_dict_keys(list(state.keys()), state)
    return state, inferred


def _maybe_prefix_backbone(state: Dict[str, torch.Tensor]) -> Dict[str, torch.Tensor]:
    """Add 'backbone.' prefix if weights come from a bare timm model (e.g., RoundTune ckpt)."""
    if any(k.startswith("backbone.") for k in state.keys()):
        return state
    return {f"backbone.{k}": v for k, v in state.items()}


def _strip_backbone_prefix(state: Dict[str, torch.Tensor]) -> Dict[str, torch.Tensor]:
    """Remove leading backbone. prefix if present."""
    if not any(k.startswith("backbone.") for k in state.keys()):
        return state
    return {k[len("backbone.") :]: v for k, v in state.items()}


def _try_load_variants(module: nn.Module, base_state: Dict[str, torch.Tensor]) -> Tuple[Dict[str, Any], Dict[str, torch.Tensor]]:
    """
    Try loading state with different prefix adjustments; pick the variant with minimal (missing+unexpected).
    Returns (info, chosen_state).
    """
    def compat_info(st: Dict[str, torch.Tensor]) -> Dict[str, Any]:
        ms = module.state_dict()
        missing = [k for k in ms.keys() if k not in st]
        unexpected = [k for k in st.keys() if k not in ms]
        shape_mismatch = [k for k in ms.keys() if (k in st and tuple(ms[k].shape) != tuple(st[k].shape))]
        return {"missing": missing, "unexpected": unexpected, "shape_mismatch": shape_mismatch}

    variants = [
        ("as_is", base_state),
        ("prefixed", _maybe_prefix_backbone(base_state)),
        ("stripped", _strip_backbone_prefix(base_state)),
    ]

    best_var = "as_is"
    best_state = base_state
    best_info = compat_info(base_state)
    best_score = len(best_info["missing"]) + len(best_info["unexpected"]) + len(best_info["shape_mismatch"])

    for var, st in variants[1:]:
        info = compat_info(st)
        score = len(info["missing"]) + len(info["unexpected"]) + len(info["shape_mismatch"])
        if score < best_score:
            best_score = score
            best_var = var
            best_state = st
            best_info = info

    return {**best_info, "variant": best_var}, best_state


def _remap_convnext_underscored(state: Dict[str, torch.Tensor]) -> Dict[str, torch.Tensor]:
    """
    Handle ConvNeXt checkpoints saved with underscores (e.g., stem_0 -> stem.0, stages_0 -> stages.0).
    This occurs when the source model used module names with underscores; timm ConvNeXt uses dots.
    """
    out: Dict[str, torch.Tensor] = {}
    for k, v in state.items():
        nk = k
        if "stem_" in nk or "stages_" in nk:
            nk = nk.replace("stem_", "stem.").replace("stages_", "stages.")
        out[nk] = v
    return out


def safe_load(module: nn.Module, state: Dict[str, torch.Tensor], strict: bool = False) -> Dict[str, Any]:
    missing, unexpected = module.load_state_dict(state, strict=strict)
    return {"missing": list(missing), "unexpected": list(unexpected)}


def drop_lora_alpha_for_mismatched_adapters(
    state: Dict[str, torch.Tensor],
    shape_mismatch: Sequence[str],
) -> Tuple[Dict[str, torch.Tensor], List[str]]:
    """Drop LoRA alpha buffers when their A/B adapter tensors cannot be loaded.

    `lora_alpha` is scalar, so it can be shape-compatible across ranks even when
    `lora_A`/`lora_B` are not. Loading it in that case silently changes the
    configured alpha for fresh adapters.
    """
    prefixes: Set[str] = set()
    for key in shape_mismatch:
        if key.endswith(".lora_A"):
            prefixes.add(key[: -len(".lora_A")])
        elif key.endswith(".lora_B"):
            prefixes.add(key[: -len(".lora_B")])

    if not prefixes:
        return state, []

    alpha_keys = sorted(f"{prefix}.lora_alpha" for prefix in prefixes if f"{prefix}.lora_alpha" in state)
    if not alpha_keys:
        return state, []

    drop = set(alpha_keys)
    filtered = {k: v for k, v in state.items() if k not in drop}
    return filtered, alpha_keys


# -----------------------------
# Logging helpers
# -----------------------------
_HISTORY_WRITE_WARNED: Set[str] = set()

def init_history_file(history_csv: Path, fieldnames: List[str]) -> None:
    ensure_dir(history_csv.parent)
    if history_csv.exists():
        try:
            with history_csv.open("r", encoding="utf-8") as f:
                header = f.readline().strip("\r\n")
            if header:
                existing = next(csv.reader([header]))
                if existing != fieldnames:
                    print(
                        f"[WARN] history.csv header mismatch for {history_csv}. "
                        f"Existing={existing} New={fieldnames}. "
                        f"Delete the file to regenerate a new header for plotting.",
                        flush=True,
                    )
        except Exception:
            pass
        return
    with history_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()


def append_history_row(history_csv: Path, row: Dict[str, Any], fieldnames: List[str]) -> None:
    try:
        with history_csv.open("a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    except PermissionError as e:
        k = str(history_csv)
        if k not in _HISTORY_WRITE_WARNED:
            _HISTORY_WRITE_WARNED.add(k)
            print(
                f"[WARN] Cannot write history csv (PermissionError): {history_csv} ({e}). "
                f"Close any app holding the file (Excel/Preview) or change --output_root.",
                flush=True,
            )
    except OSError as e:
        k = str(history_csv)
        if k not in _HISTORY_WRITE_WARNED:
            _HISTORY_WRITE_WARNED.add(k)
            print(f"[WARN] Cannot write history csv: {history_csv} ({e}).", flush=True)


def _json_default(o: Any):
    """Make common non-JSON types serializable (Path, numpy scalars, torch device/dtype, etc.)."""
    try:
        from pathlib import Path as _Path
        if isinstance(o, _Path):
            return str(o)
    except Exception:
        pass
    try:
        import numpy as _np
        if isinstance(o, (_np.integer,)):
            return int(o)
        if isinstance(o, (_np.floating,)):
            return float(o)
        if isinstance(o, (_np.bool_,)):
            return bool(o)
        if isinstance(o, _np.ndarray):
            return o.tolist()
    except Exception:
        pass
    try:
        import torch as _torch
        if isinstance(o, _torch.device):
            return str(o)
        if isinstance(o, _torch.dtype):
            return str(o)
    except Exception:
        pass
    return str(o)

def append_jsonl(jsonl_path: Path, obj: Dict[str, Any]) -> None:
    ensure_dir(jsonl_path.parent)
    with jsonl_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False, default=_json_default) + "\n")


def save_json(path: Path, obj: Dict[str, Any]) -> None:
    """Write JSON with a safe default serializer (WindowsPath, numpy, etc.)."""
    ensure_dir(path.parent)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")


def make_metrics_table_csv(out_csv: Path, rows: List[Dict[str, Any]], columns: List[str]) -> None:
    ensure_dir(out_csv.parent)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in columns})


def try_make_metrics_table_xlsx(out_xlsx: Path, rows: List[Dict[str, Any]], columns: List[str], sheet_name: str = "metrics") -> None:
    if openpyxl is None:
        return
    ensure_dir(out_xlsx.parent)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(columns)

    header_font = Font(bold=True) if Font else None
    header_fill = PatternFill("solid", fgColor="D9E1F2") if PatternFill else None
    center = Alignment(horizontal="center", vertical="center", wrap_text=True) if Alignment else None

    for j, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=j)
        if header_font:
            c.font = header_font
        if header_fill:
            c.fill = header_fill
        if center:
            c.alignment = center

    alt_fill = PatternFill("solid", fgColor="F7F7F7") if PatternFill else None
    for i, r in enumerate(rows, start=2):
        ws.append([r.get(c, "") for c in columns])
        if alt_fill and (i % 2 == 0):
            for j in range(1, len(columns) + 1):
                ws.cell(row=i, column=j).fill = alt_fill
        if center:
            for j in range(1, len(columns) + 1):
                ws.cell(row=i, column=j).alignment = center

    ws.freeze_panes = "A2"

    for j, col in enumerate(columns, 1):
        max_len = len(col)
        for i in range(2, len(rows) + 2):
            v = ws.cell(row=i, column=j).value
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = min(max_len + 2, 40)

    if "Value" in columns:
        val_col = columns.index("Value") + 1
        for i in range(2, len(rows) + 2):
            c = ws.cell(row=i, column=val_col)
            if isinstance(c.value, (float, int)):
                c.number_format = "0.000"

    wb.save(str(out_xlsx))
    wb.close()


def finalize_table(run_dir: Path, task_name: str, dataset_name: str, cfg: Dict[str, Any], best_epoch: int, best_metric: str, best_value: float, extra: Dict[str, Any]) -> None:
    rows: List[Dict[str, Any]] = []
    base = {
        "Task": task_name,
        "Dataset": dataset_name,
        "Backbone": cfg.get("backbone"),
        "ImgSize": cfg.get("img_size"),
        "BatchSize": cfg.get("batch_size"),
        "LR": cfg.get("lr"),
        "WeightDecay": cfg.get("weight_decay"),
        "BestEpoch": best_epoch,
    }
    rows.append({**base, "Metric": best_metric, "Value": float(best_value)})
    for k, v in extra.items():
        rows.append({**base, "Metric": k, "Value": float(v) if isinstance(v, (float, int, np.floating)) else v})

    cols = ["Task", "Dataset", "Backbone", "ImgSize", "BatchSize", "LR", "WeightDecay", "BestEpoch", "Metric", "Value"]
    out_csv = run_dir / "tables" / "metrics_summary.csv"
    out_xlsx = run_dir / "tables" / "metrics_summary.xlsx"
    make_metrics_table_csv(out_csv, rows, cols)
    try_make_metrics_table_xlsx(out_xlsx, rows, cols, sheet_name="metrics")


def build_writer(run_dir: Path, enabled: bool) -> Optional[Any]:
    if not enabled or SummaryWriter is None:
        return None
    tb_dir = run_dir / "tb"
    ensure_dir(tb_dir)
    return SummaryWriter(log_dir=str(tb_dir))


# -----------------------------
# DR-style grading augmentations (o_O strategy)
# -----------------------------
DR_MEAN = [108.64628601 / 255.0, 75.86886597 / 255.0, 54.34005737 / 255.0]
DR_STD = [70.53946096 / 255.0, 51.71475228 / 255.0, 43.03428563 / 255.0]
DR_U = torch.tensor(
    [
        [-0.56543481, 0.71983482, 0.40240142],
        [-0.5989477, -0.02304967, -0.80036049],
        [-0.56694071, -0.6935729, 0.44423429],
    ],
    dtype=torch.float32,
)
DR_EV = torch.tensor([1.65513492, 0.48450358, 0.1565086], dtype=torch.float32)


class KrizhevskyColorAugmentationTorch:
    """PCA-based color jitter (same as pytorch-DR-reimplement)."""
    def __init__(self, sigma: float = 0.5):
        self.sigma = sigma
        self.mean = torch.tensor([0.0], dtype=torch.float32)
        self.deviation = torch.tensor([sigma], dtype=torch.float32)

    def __call__(self, img: torch.Tensor) -> torch.Tensor:
        if self.sigma <= 0:
            return img
        if not torch.is_tensor(img):
            raise TypeError("KrizhevskyColorAugmentationTorch expects a torch.Tensor image.")
        color_vec = torch.distributions.Normal(self.mean, self.deviation).sample((3,)).squeeze()
        alpha = color_vec.to(img.device) * DR_EV.to(img.device)
        noise = torch.matmul(DR_U.to(img.device), alpha.unsqueeze(1)).view(3, 1, 1)
        return img + noise


def build_dr_grade_transform(img_size: int, is_train: bool, sigma: float = 0.5) -> Any:
    translate = 40.0 / float(max(1, img_size))
    trans: List[Any] = [T.ToPILImage()]
    if is_train:
        trans.extend([
            T.RandomResizedCrop(size=img_size, scale=(1 / 1.15, 1.15), ratio=(0.7561, 1.3225)),
            T.RandomAffine(degrees=180, translate=(translate, translate)),
            T.RandomHorizontalFlip(),
            T.RandomVerticalFlip(),
            T.ToTensor(),
            T.Normalize(mean=DR_MEAN, std=DR_STD),
            KrizhevskyColorAugmentationTorch(sigma=sigma),
        ])
    else:
        trans.extend([
            T.Resize((img_size, img_size)),
            T.ToTensor(),
            T.Normalize(mean=DR_MEAN, std=DR_STD),
        ])
    return T.Compose(trans)


# -----------------------------
# Augmentations (CFP-friendly)
# -----------------------------

# -----------------------------
# Albumentations compatibility helpers (v1.x vs v2.x)
# Albumentations v2 switched many params:
#   - RandomResizedCrop now expects `size` instead of (height, width)
#   - PadIfNeeded / ShiftScaleRotate use `fill`/`fill_mask` instead of `value`/`mask_value`
def _alb_pad_if_needed(img_size: int):
    if A is None or cv2 is None:
        raise ImportError("albumentations/opencv-python required")
    base = dict(min_height=img_size, min_width=img_size, border_mode=cv2.BORDER_CONSTANT, p=1.0)
    try:
        # Albumentations v2+
        return A.PadIfNeeded(**base, fill=0, fill_mask=0)
    except Exception:
        # Albumentations v1.x
        try:
            return A.PadIfNeeded(**base, value=0, mask_value=0)
        except Exception:
            return A.PadIfNeeded(**base)

def _alb_random_resized_crop(img_size: int, scale=(0.85, 1.0), ratio=(0.9, 1.1), p: float = 1.0):
    if A is None:
        raise ImportError("albumentations required")
    if cv2 is None:
        raise ImportError("opencv-python required")
    interp = cv2.INTER_LINEAR
    mask_interp = cv2.INTER_NEAREST
    try:
        # Albumentations v1.x
        try:
            return A.RandomResizedCrop(height=img_size, width=img_size, scale=scale, ratio=ratio, interpolation=interp, mask_interpolation=mask_interp, p=p)
        except Exception:
            return A.RandomResizedCrop(height=img_size, width=img_size, scale=scale, ratio=ratio, interpolation=interp, p=p)
    except Exception:
        # Albumentations v2+
        try:
            return A.RandomResizedCrop(size=(img_size, img_size), scale=scale, ratio=ratio, interpolation=interp, mask_interpolation=mask_interp, p=p)
        except Exception:
            return A.RandomResizedCrop(size=(img_size, img_size), scale=scale, ratio=ratio, interpolation=interp, p=p)

def _alb_shift_scale_rotate(*, shift_limit=0.02, scale_limit=0.08, rotate_limit=10, p: float = 0.5):
    if A is None or cv2 is None:
        raise ImportError("albumentations/opencv-python required")
    base = dict(shift_limit=shift_limit, scale_limit=scale_limit, rotate_limit=rotate_limit,
                border_mode=cv2.BORDER_CONSTANT, p=p)
    base_interp = dict(interpolation=cv2.INTER_LINEAR)
    base_mask_interp = dict(mask_interpolation=cv2.INTER_NEAREST)
    # Albumentations v2+ uses fill/fill_mask; v1.x used value/mask_value.
    try:
        try:
            return A.ShiftScaleRotate(**base, **base_interp, **base_mask_interp, fill=0, fill_mask=0)
        except Exception:
            return A.ShiftScaleRotate(**base, **base_interp, fill=0, fill_mask=0)
    except Exception:
        try:
            try:
                return A.ShiftScaleRotate(**base, **base_interp, **base_mask_interp, value=0, mask_value=0)
            except Exception:
                return A.ShiftScaleRotate(**base, **base_interp, value=0, mask_value=0)
        except Exception:
            try:
                return A.ShiftScaleRotate(**base, **base_interp, **base_mask_interp)
            except Exception:
                return A.ShiftScaleRotate(**base)

def _alb_grid_distortion(*, num_steps: int = 5, distort_limit: float = 0.03, p: float = 0.2):
    if A is None or cv2 is None:
        raise ImportError("albumentations/opencv-python required")
    base = dict(num_steps=int(num_steps), distort_limit=float(distort_limit), border_mode=cv2.BORDER_CONSTANT, p=float(p))
    base_interp = dict(interpolation=cv2.INTER_LINEAR)
    base_mask_interp = dict(mask_interpolation=cv2.INTER_NEAREST)
    # Albumentations v2+ uses fill/fill_mask; v1.x used value/mask_value.
    try:
        try:
            return A.GridDistortion(**base, **base_interp, **base_mask_interp, fill=0, fill_mask=0)
        except Exception:
            return A.GridDistortion(**base, **base_interp, fill=0, fill_mask=0)
    except Exception:
        try:
            try:
                return A.GridDistortion(**base, **base_interp, **base_mask_interp, value=0, mask_value=0)
            except Exception:
                return A.GridDistortion(**base, **base_interp, value=0, mask_value=0)
        except Exception:
            try:
                return A.GridDistortion(**base, **base_interp, **base_mask_interp)
            except Exception:
                return A.GridDistortion(**base)

def _alb_longest_max_size(img_size: int):
    if A is None or cv2 is None:
        raise ImportError("albumentations/opencv-python required")
    base = dict(max_size=img_size)
    try:
        return A.LongestMaxSize(**base, interpolation=cv2.INTER_LINEAR, mask_interpolation=cv2.INTER_NEAREST)
    except Exception:
        try:
            return A.LongestMaxSize(**base, interpolation=cv2.INTER_LINEAR)
        except Exception:
            return A.LongestMaxSize(**base)


def build_augs(task: str, img_size: int):
    """Build augmentations for the public quality/grade training entry points."""
    if A is None or ToTensorV2 is None or cv2 is None:
        # Fallback torchvision path used by QualityDataset when Albumentations is absent.
        return None

    task = str(task).lower().strip()
    if task == "qual":
        return A.Compose([
            _alb_random_resized_crop(img_size, scale=(0.8, 1.0), ratio=(0.9, 1.1), p=1.0),
            A.HorizontalFlip(p=0.5),
            A.RandomBrightnessContrast(p=0.35),
            A.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),
            ToTensorV2(),
        ])

    if task == "grade":
        return A.Compose([
            _alb_longest_max_size(img_size),
            _alb_pad_if_needed(img_size, img_size),
            A.HorizontalFlip(p=0.5),
            _alb_shift_scale_rotate(shift_limit=0.03, scale_limit=0.08, rotate_limit=12, p=0.6),
            A.RandomBrightnessContrast(p=0.35),
            A.HueSaturationValue(p=0.25),
            A.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),
            ToTensorV2(),
        ])

    raise ValueError(f"Unsupported task for public augmentations: {task!r}")

def build_val_aug(img_size: int) -> Any:
    if A is None:
        raise ImportError("albumentations required")
    if cv2 is None:
        raise ImportError("opencv-python required")
    norm = A.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225))
    return A.Compose([
        _alb_longest_max_size(img_size),
        _alb_pad_if_needed(img_size),
        A.CenterCrop(height=img_size, width=img_size),
        norm,
        ToTensorV2(),
    ])


# -----------------------------
# Datasets (loss masking ready)
# -----------------------------
def get_first_existing_key(keys: List[str], candidates: List[str]) -> Optional[str]:
    ks = set(keys)
    for c in candidates:
        if c in ks:
            return c
    return None


class QualityDataset(Dataset):
    """
    Expected CSV columns:
      - image_path (required)
      - quality label column (one of): quality, overall, gradability, label_quality, q
    Labels can be missing; missing -> label = -1 (loss masked).
    """
    def __init__(self, rows: List[Dict[str, str]], project_root: Path, img_size: int, is_train: bool, collapse_to_binary: bool = False):
        self.rows = rows
        self.project_root = project_root
        self.aug = build_augs("qual", img_size) if is_train else build_val_aug(img_size)

        self.collapse_to_binary = bool(collapse_to_binary)

        keys = list(rows[0].keys())
        self.label_key = get_first_existing_key(keys, ["label", "quality", "overall", "gradability", "label_quality", "q"])
        if self.label_key is None:
            raise ValueError(f"Quality CSV missing a quality label column. Found keys: {sorted(keys)}")

    def __len__(self) -> int:
        return len(self.rows)

    def __getitem__(self, idx: int) -> Dict[str, Any]:
        r = self.rows[idx]
        img_path = resolve_path(self.project_root, r["image_path"])
        img = cv2.imread(img_path, cv2.IMREAD_COLOR)
        if img is None:
            raise FileNotFoundError(img_path)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        lab_raw = r.get(self.label_key, "")
        if lab_raw == "" or lab_raw is None:
            y = -1
        else:
            y = int(float(lab_raw))

        # Optional: collapse multi-class quality labels to binary gradable/ungradable (for more stable training).
        # If labels are already 0/1, keep unchanged.
        if y >= 0 and self.collapse_to_binary:
            if y > 1:
                # Common 3-class convention: 0=Good, 1=Usable, 2=Reject -> gradable={0,1} -> 1
                y = 1 if y in (0, 1) else 0

        out = self.aug(image=img)
        return {"image": out["image"], "label": torch.tensor(y, dtype=torch.long), "has_label": torch.tensor(int(y >= 0)), "path": img_path}


class GradeDataset(Dataset):
    """
    Expected CSV columns:
      - image_path (required)
      - grade label column: label / grade / dr_grade / icdr
    Missing label -> -1 (loss masked).
    """
    def __init__(self, rows: List[Dict[str, str]], project_root: Path, img_size: int, is_train: bool):
        self.rows = rows
        self.project_root = project_root
        self.transform = build_dr_grade_transform(img_size, is_train)

        keys = list(rows[0].keys())
        self.label_key = get_first_existing_key(keys, ["label", "grade", "dr_grade", "icdr"])
        if self.label_key is None:
            raise ValueError(f"Grade CSV missing label column. Found keys: {sorted(keys)}")

    def __len__(self) -> int:
        return len(self.rows)

    def __getitem__(self, idx: int) -> Dict[str, Any]:
        r = self.rows[idx]
        img_path = resolve_path(self.project_root, r["image_path"])
        img = cv2.imread(img_path, cv2.IMREAD_COLOR)
        if img is None:
            raise FileNotFoundError(img_path)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        lab_raw = r.get(self.label_key, "")
        if lab_raw == "" or lab_raw is None:
            y = -1
        else:
            y = int(float(lab_raw))

        img_t = self.transform(img)
        return {"image": img_t, "label": torch.tensor(y, dtype=torch.long), "has_label": torch.tensor(int(y >= 0)), "path": img_path}

class GradeAlbDataset(Dataset):
    """
    Grade dataset with Albumentations + ImageNet normalization (better matches ImageNet-pretrained backbones).
    Expected CSV columns:
      - image_path (required)
      - grade label column: label / grade / dr_grade / icdr
    Missing label -> -1 (loss masked).
    """
    def __init__(self, rows: List[Dict[str, str]], project_root: Path, img_size: int, is_train: bool):
        self.rows = rows
        self.project_root = project_root
        self.aug = build_augs("grade", img_size) if is_train else build_val_aug(img_size)

        keys = list(rows[0].keys())
        self.label_key = get_first_existing_key(keys, ["label", "grade", "dr_grade", "icdr"])
        if self.label_key is None:
            raise ValueError(f"Grade CSV missing label column. Found keys: {sorted(keys)}")

    def __len__(self) -> int:
        return len(self.rows)

    def __getitem__(self, idx: int) -> Dict[str, Any]:
        r = self.rows[idx]
        img_path = resolve_path(self.project_root, r["image_path"])
        img = cv2.imread(img_path, cv2.IMREAD_COLOR)
        if img is None:
            raise FileNotFoundError(img_path)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        lab_raw = r.get(self.label_key, "")
        if lab_raw == "" or lab_raw is None:
            y = -1
        else:
            y = int(float(lab_raw))

        out = self.aug(image=img)
        return {"image": out["image"], "label": torch.tensor(y, dtype=torch.long), "has_label": torch.tensor(int(y >= 0)), "path": img_path}






# -----------------------------
# Models
# -----------------------------
class PatchEmbed(nn.Module):
    """2D image to patch embedding (Conv2d)."""
    def __init__(self, img_size: int = 224, patch_size: int = 16, in_chans: int = 3, embed_dim: int = 1024):
        super().__init__()
        self.img_size = img_size
        self.patch_size = patch_size
        self.grid_size = (img_size // patch_size, img_size // patch_size)
        self.num_patches = self.grid_size[0] * self.grid_size[1]
        self.proj = nn.Conv2d(in_chans, embed_dim, kernel_size=patch_size, stride=patch_size)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x = self.proj(x)  # B, C, H/ps, W/ps
        x = x.flatten(2).transpose(1, 2)  # B, N, C
        return x


class Attention(nn.Module):
    def __init__(self, dim: int, num_heads: int = 16):
        super().__init__()
        self.num_heads = num_heads
        head_dim = dim // num_heads
        self.scale = head_dim ** -0.5
        self.qkv = nn.Linear(dim, dim * 3, bias=True)
        self.proj = nn.Linear(dim, dim)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        B, N, C = x.shape
        qkv = self.qkv(x).reshape(B, N, 3, self.num_heads, C // self.num_heads).permute(2, 0, 3, 1, 4)
        q, k, v = qkv[0], qkv[1], qkv[2]
        attn = (q @ k.transpose(-2, -1)) * self.scale
        attn = attn.softmax(dim=-1)
        out = (attn @ v).transpose(1, 2).reshape(B, N, C)
        out = self.proj(out)
        return out


class LoRALinear(nn.Module):
    """
    Minimal LoRA wrapper for nn.Linear.
    Forward: base(x) + (alpha/r) * (dropout(x) @ A^T @ B^T)

    - base weights stay frozen in LoRA tuning modes.
    - lora_alpha is stored as a buffer so it is saved/loaded with the state_dict.
    """
    def __init__(self, base: nn.Linear, *, r: int, alpha: float = 1.0, dropout: float = 0.0):
        super().__init__()
        if not isinstance(base, nn.Linear):
            raise TypeError("LoRALinear expects an nn.Linear base module.")
        if int(r) <= 0:
            raise ValueError("LoRALinear requires r>0.")
        self.base = base
        self.r = int(r)
        dev = base.weight.device
        self.register_buffer("lora_alpha", torch.tensor(float(alpha), dtype=torch.float32, device=dev))
        self.lora_dropout = nn.Dropout(float(dropout)) if float(dropout) > 0 else nn.Identity()

        # LoRA modules may be injected after the model is already on GPU; keep params on the same device.
        self.lora_A = nn.Parameter(torch.empty(self.r, base.in_features, device=dev, dtype=base.weight.dtype))
        self.lora_B = nn.Parameter(torch.zeros(base.out_features, self.r, device=dev, dtype=base.weight.dtype))
        nn.init.kaiming_uniform_(self.lora_A, a=math.sqrt(5))
        nn.init.zeros_(self.lora_B)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        out = self.base(x)
        x_d = self.lora_dropout(x)
        # (.., in) -> (.., r) -> (.., out)
        xa = F.linear(x_d, self.lora_A)
        xb = F.linear(xa, self.lora_B)
        scale = (self.lora_alpha / float(self.r)).to(dtype=xb.dtype)
        return out + xb * scale


class MLP(nn.Module):
    def __init__(self, dim: int, mlp_ratio: float = 4.0):
        super().__init__()
        hidden = int(dim * mlp_ratio)
        self.fc1 = nn.Linear(dim, hidden)
        self.act = nn.GELU()
        self.fc2 = nn.Linear(hidden, dim)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return self.fc2(self.act(self.fc1(x)))


class Block(nn.Module):
    def __init__(self, dim: int, num_heads: int, mlp_ratio: float = 4.0):
        super().__init__()
        self.norm1 = nn.LayerNorm(dim)
        self.attn = Attention(dim, num_heads=num_heads)
        self.norm2 = nn.LayerNorm(dim)
        self.mlp = MLP(dim, mlp_ratio=mlp_ratio)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x = x + self.attn(self.norm1(x))
        x = x + self.mlp(self.norm2(x))
        return x


class MAEViTBackbone(nn.Module):
    """
    Minimal MAE ViT backbone (encoder + decoder params for weight loading).
    Encoder output: multi-scale feature maps + cls embedding.
    """
    def __init__(
        self,
        img_size: int = 224,
        patch_size: int = 16,
        in_chans: int = 3,
        embed_dim: int = 1024,
        depth: int = 24,
        num_heads: int = 16,
        mlp_ratio: float = 4.0,
        decoder_embed_dim: int = 512,
        decoder_depth: int = 8,
        decoder_num_heads: int = 16,
    ):
        super().__init__()
        self.img_size = img_size
        self.patch_size = patch_size
        self.embed_dim = embed_dim
        self.decoder_embed_dim = decoder_embed_dim
        self.patch_embed = PatchEmbed(img_size, patch_size, in_chans, embed_dim)
        num_patches = self.patch_embed.num_patches

        self.cls_token = nn.Parameter(torch.zeros(1, 1, embed_dim))
        self.pos_embed = nn.Parameter(torch.zeros(1, num_patches + 1, embed_dim))

        self.blocks = nn.ModuleList([Block(embed_dim, num_heads, mlp_ratio) for _ in range(depth)])
        self.norm = nn.LayerNorm(embed_dim)

        # Decoder params (not used in forward, only for weight compatibility)
        self.decoder_embed = nn.Linear(embed_dim, decoder_embed_dim, bias=True)
        self.mask_token = nn.Parameter(torch.zeros(1, 1, decoder_embed_dim))
        self.decoder_pos_embed = nn.Parameter(torch.zeros(1, num_patches + 1, decoder_embed_dim))
        self.decoder_blocks = nn.ModuleList([Block(decoder_embed_dim, decoder_num_heads, mlp_ratio) for _ in range(decoder_depth)])
        self.decoder_norm = nn.LayerNorm(decoder_embed_dim)
        self.decoder_pred = nn.Linear(decoder_embed_dim, patch_size ** 2 * in_chans, bias=True)

        self.feature_channels = [embed_dim, embed_dim, embed_dim, embed_dim]
        self._init_weights()

    def _init_weights(self):
        torch.nn.init.normal_(self.cls_token, std=0.02)
        torch.nn.init.normal_(self.mask_token, std=0.02)
        torch.nn.init.normal_(self.pos_embed, std=0.02)
        torch.nn.init.normal_(self.decoder_pos_embed, std=0.02)
        for m in self.modules():
            if isinstance(m, nn.Linear):
                torch.nn.init.xavier_uniform_(m.weight)
                if m.bias is not None:
                    nn.init.constant_(m.bias, 0)
            elif isinstance(m, nn.LayerNorm):
                nn.init.constant_(m.bias, 0)
                nn.init.constant_(m.weight, 1.0)

    def interpolate_pos_encoding(self, x: torch.Tensor, H: int, W: int) -> torch.Tensor:
        # x: B,N,C (N includes cls)
        npatch = x.shape[1] - 1
        N = self.pos_embed.shape[1] - 1
        if npatch == N and W == H == int((N) ** 0.5) * self.patch_size:
            return self.pos_embed
        class_pos_embed = self.pos_embed[:, :1, :]
        patch_pos_embed = self.pos_embed[:, 1:, :]
        dim = x.shape[-1]
        h0 = H // self.patch_size
        w0 = W // self.patch_size
        patch_pos = patch_pos_embed.reshape(1, int(N ** 0.5), int(N ** 0.5), dim).permute(0, 3, 1, 2)
        patch_pos = F.interpolate(patch_pos, size=(h0, w0), mode="bicubic", align_corners=False)
        patch_pos = patch_pos.permute(0, 2, 3, 1).reshape(1, h0 * w0, dim)
        return torch.cat((class_pos_embed, patch_pos), dim=1)

    def forward(self, x: torch.Tensor) -> Tuple[List[torch.Tensor], torch.Tensor]:
        B, _, H, W = x.shape
        x_tokens = self.patch_embed(x)
        pos = self.interpolate_pos_encoding(x_tokens, H, W)
        cls_tokens = self.cls_token.expand(B, -1, -1)
        x_tokens = torch.cat((cls_tokens, x_tokens), dim=1)
        x_tokens = x_tokens + pos

        # Use intermediate encoder blocks for a more meaningful feature hierarchy (helps UNet/DeepLab decoders).
        depth = len(self.blocks)
        i1 = max(0, depth // 4 - 1)
        i2 = max(0, depth // 2 - 1)
        i3 = max(0, (3 * depth) // 4 - 1)
        capture = {i1, i2, i3, depth - 1}
        snaps: Dict[int, torch.Tensor] = {}
        for i, blk in enumerate(self.blocks):
            x_tokens = blk(x_tokens)
            if i in capture:
                snaps[i] = x_tokens
        x_tokens = self.norm(x_tokens)

        cls_out = x_tokens[:, 0]
        h = H // self.patch_size
        w = W // self.patch_size

        def _to_fmap(tok: torch.Tensor) -> torch.Tensor:
            pt = tok[:, 1:]
            return pt.transpose(1, 2).reshape(B, self.embed_dim, h, w)

        # Build a 4-level pyramid roughly matching common CNN stage strides:
        #   f1 ~ 1/4, f2 ~ 1/8, f3 ~ 1/16 (patch map), f4 ~ 1/32
        # Each level uses a different encoder depth snapshot, then spatially rescaled.
        fmap_early = _to_fmap(self.norm(snaps.get(i1, x_tokens)))
        fmap_mid = _to_fmap(self.norm(snaps.get(i2, x_tokens)))
        fmap_late = _to_fmap(self.norm(snaps.get(i3, x_tokens)))
        fmap_final = _to_fmap(x_tokens)

        f1 = F.interpolate(fmap_early, scale_factor=4.0, mode="bilinear", align_corners=False)
        f2 = F.interpolate(fmap_mid, scale_factor=2.0, mode="bilinear", align_corners=False)
        f3 = fmap_late
        f4 = F.interpolate(fmap_final, scale_factor=0.5, mode="bilinear", align_corners=False)
        feats = [f1, f2, f3, f4]
        return feats, cls_out


class SharedBackbone(nn.Module):
    """Backbone wrapper: returns multi-scale feature maps + global pooled embedding."""
    def __init__(
        self,
        backbone_name: str,
        *,
        pretrained: bool = False,
        timm_pretrained_file: Optional[Path] = None,
    ):
        super().__init__()
        backbone_name = backbone_name.lower()
        self.is_timm_backbone = False
        if backbone_name in ("mae_vit_large_patch16_224", "mae_vit_large_patch16"):
            self.backbone = MAEViTBackbone(
                img_size=224,
                patch_size=16,
                embed_dim=1024,
                depth=24,
                num_heads=16,
                mlp_ratio=4.0,
                decoder_embed_dim=512,
                decoder_depth=8,
                decoder_num_heads=16,
            )
        elif backbone_name in ("mae_vit_base_patch16_224", "mae_vit_base_patch16"):
            self.backbone = MAEViTBackbone(
                img_size=224,
                patch_size=16,
                embed_dim=768,
                depth=12,
                num_heads=12,
                mlp_ratio=4.0,
                decoder_embed_dim=512,
                decoder_depth=8,
                decoder_num_heads=16,
            )
        else:
            if timm is None:
                raise ImportError("timm is required for non-MAE backbones.")
            timm_global_pool = "token" if backbone_name.startswith(("vit_", "deit_")) else "avg"
            timm_kwargs: Dict[str, Any] = {}
            if timm_pretrained_file is not None:
                timm_pretrained_file = Path(timm_pretrained_file)
                if not timm_pretrained_file.exists():
                    raise FileNotFoundError(f"timm pretrained file not found: {timm_pretrained_file}")
                timm_kwargs["pretrained_cfg_overlay"] = {
                    "file": str(timm_pretrained_file),
                    "hf_hub_id": None,
                    "url": "",
                }
            self.backbone = timm.create_model(
                backbone_name,
                pretrained=pretrained,
                num_classes=0,
                global_pool=timm_global_pool,
                **timm_kwargs,
            )
            self.is_timm_backbone = True
            num_features = int(getattr(self.backbone, "num_features", 0) or 0)
            if num_features <= 0:
                raise ValueError(f"Unable to determine num_features for timm backbone '{backbone_name}'.")
            self.backbone.embed_dim = num_features
            self.feature_channels = [num_features, num_features, num_features, num_features]
        if not self.is_timm_backbone:
            self.feature_channels = self.backbone.feature_channels

    def forward(self, x: torch.Tensor) -> Tuple[List[torch.Tensor], torch.Tensor]:
        if self.is_timm_backbone:
            z = self.backbone(x)
            if z.ndim > 2:
                z = z.flatten(2).mean(dim=-1)
            fmap = z[:, :, None, None]
            feats = [fmap, fmap, fmap, fmap]
            return feats, z
        feats, z = self.backbone(x)
        return feats, z


    def __init__(self, backbone: SharedBackbone, num_classes: int, mask_mode: str, decoder_name: str = "fpn", img_size_hint: Optional[int] = None):
        super().__init__()
        self.backbone = backbone
        self.decoder_name = str(decoder_name).lower().strip()

        if self.decoder_name in ("fpn", "fpn_gn2", "fpn_head2", ""):
            self.decoder = FPNDecoder(backbone.feature_channels, fpn_dim=256)
            decoder_out = 256
        elif self.decoder_name in ("unet", "u-net"):
            self.decoder = UNetDecoder(backbone.feature_channels, out_ch=256)
            decoder_out = 256
        elif self.decoder_name in ("deeplabv3p", "deeplabv3+", "deeplab", "deeplabv3plus"):
            hint = int(img_size_hint) if (img_size_hint is not None and int(img_size_hint) > 0) else 512
            atrous_rates = _deeplab_aspp_rates_for_img_size(hint)
            self.decoder = DeepLabV3PlusDecoder(backbone.feature_channels, out_ch=256, low_ch=48, atrous_rates=atrous_rates)
            decoder_out = 256
        else:
            raise ValueError(f"Unsupported --seg_decoder '{decoder_name}'. Choose from: unet, deeplabv3p, fpn, fpn_gn2")
        self.mask_mode = mask_mode
        if self.decoder_name in ("fpn_gn2", "fpn_head2"):
            self.head = SegHeadGN2(decoder_out, num_classes)
        else:
            self.head = nn.Conv2d(decoder_out, num_classes, 1)

    def forward(self, x: torch.Tensor) -> Tuple[torch.Tensor, torch.Tensor]:
        feats, z = self.backbone(x)
        dec = self.decoder(feats)
        logits = self.head(dec)
        logits = F.interpolate(logits, size=x.shape[-2:], mode="bilinear", align_corners=False)
        return logits, z


class QualityModel(nn.Module):
    def __init__(self, backbone: SharedBackbone, num_classes_quality: int = 3, dropout: float = 0.2):
        super().__init__()
        self.backbone = backbone
        dim = backbone.feature_channels[-1]
        self.dropout = nn.Dropout(dropout)
        self.head = nn.Linear(dim, num_classes_quality)

    def forward(self, x: torch.Tensor) -> Tuple[torch.Tensor, torch.Tensor]:
        _, z = self.backbone(x)
        logits = self.head(self.dropout(z))
        return logits, z




class GradeModel(nn.Module):
    """Image-only grading (ICDR 0鈥?)."""
    def __init__(self, backbone: SharedBackbone, num_classes: int = 5, dropout: float = 0.2):
        super().__init__()
        self.backbone = backbone
        dim = backbone.feature_channels[-1]
        self.dropout = nn.Dropout(dropout)
        self.head = nn.Linear(dim, num_classes)

    def forward(self, x: torch.Tensor) -> Tuple[torch.Tensor, torch.Tensor]:
        _, z = self.backbone(x)
        logits = self.head(self.dropout(z))
        return logits, z




# -----------------------------
# Losses / Metrics
# -----------------------------


@torch.no_grad()
def cls_metrics(logits: torch.Tensor, y: torch.Tensor) -> Dict[str, float]:
    pred = torch.argmax(logits, dim=1).cpu().numpy()
    y_np = y.cpu().numpy()
    out: Dict[str, float] = {}
    out["acc"] = float((pred == y_np).mean())
    if f1_score is not None:
        out["macro_f1"] = float(f1_score(y_np, pred, average="macro"))
    else:
        out["macro_f1"] = macro_f1_numpy(y_np, pred, int(logits.shape[1]))
    if cohen_kappa_score is not None and len(np.unique(y_np)) > 1:
        out["qwk"] = float(cohen_kappa_score(y_np, pred, weights="quadratic"))
    return out


@torch.no_grad()
def cls_metrics_from_probs(probs: torch.Tensor, y: torch.Tensor) -> Dict[str, float]:
    """Classification metrics from class probabilities (B,C)."""
    pred = torch.argmax(probs, dim=1).cpu().numpy()
    y_np = y.cpu().numpy()
    out: Dict[str, float] = {}
    out["acc"] = float((pred == y_np).mean())
    if f1_score is not None:
        out["macro_f1"] = float(f1_score(y_np, pred, average="macro", zero_division=0))
    else:
        out["macro_f1"] = macro_f1_numpy(y_np, pred, int(probs.shape[1]))
    if cohen_kappa_score is not None and len(np.unique(y_np)) > 1:
        out["qwk"] = float(cohen_kappa_score(y_np, pred, weights="quadratic"))
    return out


def _edl_evidence(logits: torch.Tensor, mode: str = "relu") -> torch.Tensor:
    m = str(mode).lower().strip()
    if m in ("relu", "re" ):
        return F.relu(logits)
    if m in ("softplus", "sp"):
        return F.softplus(logits)
    raise ValueError(f"Unsupported edl_evidence='{mode}'. Choose from: relu, softplus")


def edl_dirichlet_probs(logits: torch.Tensor, *, evidence: str = "relu") -> Tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
    """Return (probs, alpha, S) from raw logits using Dirichlet evidence."""
    e = _edl_evidence(logits, mode=evidence)
    alpha = e + 1.0
    S = alpha.sum(dim=1, keepdim=True).clamp(min=1e-6)
    probs = alpha / S
    return probs, alpha, S


def edl_kl_divergence_to_uniform(alpha: torch.Tensor) -> torch.Tensor:
    """KL(Dir(alpha) || Dir(1)) for each sample. alpha: (B,K), alpha>=1."""
    alpha = alpha.to(torch.float32)
    K = int(alpha.shape[1])
    ones = torch.ones((1, K), dtype=alpha.dtype, device=alpha.device)
    sum_alpha = alpha.sum(dim=1, keepdim=True)
    sum_ones = ones.sum(dim=1, keepdim=True)

    # log B(alpha) - log B(ones)
    first = (
        torch.lgamma(sum_alpha)
        - torch.lgamma(alpha).sum(dim=1, keepdim=True)
        - torch.lgamma(sum_ones)
        + torch.lgamma(ones).sum(dim=1, keepdim=True)
    )
    # (alpha-ones) * (digamma(alpha) - digamma(sum_alpha))
    second = ((alpha - ones) * (torch.digamma(alpha) - torch.digamma(sum_alpha))).sum(dim=1, keepdim=True)
    kl = first + second
    return kl.view(-1)


def edl_digamma_loss(
    logits: torch.Tensor,
    target: torch.Tensor,
    *,
    num_classes: int = 5,
    evidence: str = "relu",
    weight: Optional[torch.Tensor] = None,
    kl_weight: float = 1.0,
    anneal: float = 1.0,
) -> torch.Tensor:
    """EDL loss = digamma data term + annealed KL(Dir(alpha_tilde)||Dir(1))."""
    logits = logits.to(torch.float32)
    target = target.clamp(0, int(num_classes) - 1).long()
    probs, alpha, S = edl_dirichlet_probs(logits, evidence=evidence)
    y_oh = F.one_hot(target, num_classes=int(num_classes)).to(dtype=logits.dtype)
    # data term: sum_k y_k (psi(S) - psi(alpha_k))
    A = (y_oh * (torch.digamma(S) - torch.digamma(alpha))).sum(dim=1)  # (B,)

    # remove evidence for true class from KL (discourage misleading evidence)
    alpha_tilde = (alpha - 1.0) * (1.0 - y_oh) + 1.0
    kl = edl_kl_divergence_to_uniform(alpha_tilde)  # (B,)

    sample_loss = A + float(kl_weight) * float(anneal) * kl

    if weight is not None:
        w = weight[target].to(dtype=sample_loss.dtype, device=sample_loss.device)
        denom = w.sum().clamp(min=1.0)
        return (sample_loss * w).sum() / denom

    return sample_loss.mean()


def focal_loss_logits(
    logits: torch.Tensor,
    target: torch.Tensor,
    *,
    gamma: float = 2.0,
    weight: Optional[torch.Tensor] = None,
    reduction: str = "mean",
) -> torch.Tensor:
    """
    Standard focal loss for multi-class classification using logits.
    Uses cross-entropy as the base loss and applies (1 - p_t)^gamma modulation.
    """
    ce = F.cross_entropy(logits, target, weight=weight, reduction="none")
    pt = torch.exp(-ce)
    loss = ((1.0 - pt) ** float(gamma)) * ce
    if reduction == "mean":
        return loss.mean()
    if reduction == "sum":
        return loss.sum()
    if reduction == "none":
        return loss
    raise ValueError(f"Unsupported reduction='{reduction}'")


# ==============================================================================
# 鍒涙柊鎹熷け鍑芥暟 (Innovation Loss Functions)
# ==============================================================================

def corn_loss(
    logits: torch.Tensor,
    target: torch.Tensor,
    num_classes: int = 5,
    weight: Optional[torch.Tensor] = None,
) -> torch.Tensor:
    """
    CORN Loss (Conditional Ordinal Regression for Neural Networks)
    
    灏嗘湁搴忓垎绫婚棶棰樿浆鎹负K-1涓簩鍒嗙被闂:
    - 浠诲姟1: P(y > 0)
    - 浠诲姟2: P(y > 1) 
    - 浠诲姟3: P(y > 2)
    - 浠诲姟4: P(y > 3)
    
    Args:
        logits: (B, num_classes) 鎴?(B, num_classes-1) 鐨勬ā鍨嬭緭鍑?        target: (B,) 鐪熷疄鏍囩 0-4
        num_classes: 绫诲埆鏁伴噺 (榛樿5)
        weight: 鍙€夌殑绫诲埆鏉冮噸
    
    Returns:
        鏍囬噺鎹熷け鍊?    """
    batch_size = logits.shape[0]
    
    # If logits are (B, num_classes), keep the first num_classes-1 outputs for CORN.
    if logits.shape[1] == num_classes:
        logits = logits[:, :-1]  # (B, num_classes-1)
    
    losses = []
    for k in range(num_classes - 1):
        # 绗琸涓换鍔? 棰勬祴 y > k
        binary_target = (target > k).float()  # (B,)
        logit_k = logits[:, k]  # (B,)
        
        # 璁＄畻浜屽垎绫讳氦鍙夌喌
        loss_k = F.binary_cross_entropy_with_logits(logit_k, binary_target, reduction="none")
        
        # 鍙€? 搴旂敤绫诲埆鏉冮噸
        if weight is not None:
            # 鏍规嵁鍘熷鏍囩鍔犳潈
            sample_weight = weight[target.clamp(0, num_classes - 1)]
            loss_k = loss_k * sample_weight
        
        losses.append(loss_k.mean())
    
    return torch.stack(losses).mean()


def corn_predict(logits: torch.Tensor, num_classes: int = 5) -> torch.Tensor:
    """
    浠嶤ORN妯″瀷杈撳嚭杞崲涓虹被鍒娴?    
    Args:
        logits: (B, num_classes) 鎴?(B, num_classes-1)
        num_classes: 绫诲埆鏁伴噺
    
    Returns:
        (B,) 棰勬祴绫诲埆 0 鍒?num_classes-1
    """
    if logits.shape[1] == num_classes:
        logits = logits[:, :-1]  # (B, num_classes-1)
    
    # 璁＄畻姣忎釜闃堝€肩殑姒傜巼
    probs = torch.sigmoid(logits)  # (B, num_classes-1)
    
    # 棰勬祴绫诲埆 = 婊¤冻 P(y > k) > 0.5 鐨勬渶澶?k+1
    predictions = (probs > 0.5).sum(dim=1)  # (B,)
    
    return predictions.long()


@torch.no_grad()
def cls_metrics_corn(logits: torch.Tensor, y: torch.Tensor, num_classes: int = 5) -> Dict[str, float]:
    """CORN涓撶敤鐨刴etrics鍑芥暟锛屼娇鐢╟orn_predict杩涜棰勬祴"""
    pred = corn_predict(logits, num_classes=num_classes).cpu().numpy()
    y_np = y.cpu().numpy()
    out: Dict[str, float] = {}
    out["acc"] = float((pred == y_np).mean())
    if f1_score is not None:
        out["macro_f1"] = float(f1_score(y_np, pred, average="macro", zero_division=0))
    else:
        out["macro_f1"] = macro_f1_numpy(y_np, pred, num_classes)
    if cohen_kappa_score is not None and len(np.unique(y_np)) > 1:
        out["qwk"] = float(cohen_kappa_score(y_np, pred, weights="quadratic"))
    return out


def soft_qwk_loss(
    logits: torch.Tensor,
    target: torch.Tensor,
    num_classes: int = 5,
    eps: float = 1e-6,
) -> torch.Tensor:
    """
    鍙井鍒嗙殑 Quadratic Weighted Kappa Loss
    
    浣跨敤杞爣绛捐绠?QWK锛屼娇鍏跺彲寰垎鐢ㄤ簬鍙嶅悜浼犳挱
    
    Args:
        logits: (B, num_classes) 妯″瀷杈撳嚭
        target: (B,) 鐪熷疄鏍囩
        num_classes: 绫诲埆鏁伴噺
        eps: 鏁板€肩ǔ瀹氭€?    
    Returns:
        1 - QWK (鏈€灏忓寲姝ゅ€兼潵鏈€澶у寲 QWK)
    """
    device = logits.device
    batch_size = logits.shape[0]
    
    # Soft prediction distribution.
    probs = F.softmax(logits, dim=1)  # (B, C)
    
    # 鐪熷疄鏍囩鐨?one-hot
    target_clamped = target.clamp(0, num_classes - 1)
    targets_oh = F.one_hot(target_clamped, num_classes).float()  # (B, C)
    
    # 棰勮绠楁潈閲嶇煩闃?W[i,j] = (i-j)^2 / (num_classes-1)^2
    i_idx = torch.arange(num_classes, device=device).float().view(-1, 1)
    j_idx = torch.arange(num_classes, device=device).float().view(1, -1)
    weights = ((i_idx - j_idx) ** 2) / ((num_classes - 1) ** 2)  # (C, C)
    
    # 杞贩娣嗙煩闃?O[i,j] = sum_b P(pred=j|b) * 1(true=i|b)
    O = torch.matmul(targets_oh.T, probs)  # (C, C)
    O = O / (O.sum() + eps)
    
    # 鏈熸湜娣锋穯鐭╅樀 E[i,j] = P(true=i) * P(pred=j)
    hist_true = targets_oh.sum(dim=0)  # (C,)
    hist_pred = probs.sum(dim=0)       # (C,)
    E = torch.outer(hist_true, hist_pred)
    E = E / (E.sum() + eps)
    
    # QWK = 1 - sum(W*O) / sum(W*E)
    num = (weights * O).sum()
    den = (weights * E).sum()
    
    qwk = 1.0 - num / (den + eps)
    
    # 杩斿洖 loss = 1 - QWK (鏈€灏忓寲姝ゅ€兼潵鏈€澶у寲 QWK)
    return 1.0 - qwk


def combined_ce_qwk_loss(
    logits: torch.Tensor,
    target: torch.Tensor,
    num_classes: int = 5,
    qwk_weight: float = 0.5,
    weight: Optional[torch.Tensor] = None,
    label_smoothing: float = 0.1,
) -> torch.Tensor:
    """
    缁勫悎鎹熷け: CE + QWK
    
    CE 鎻愪緵绋冲畾鐨勬搴︼紝QWK 鐩存帴浼樺寲鐩爣鎸囨爣
    
    Args:
        logits: (B, num_classes) 妯″瀷杈撳嚭
        target: (B,) 鐪熷疄鏍囩
        num_classes: 绫诲埆鏁伴噺
        qwk_weight: QWK 鎹熷け鐨勬潈閲?(0-1)
        weight: CE 鐨勭被鍒潈閲?        label_smoothing: CE 鐨勬爣绛惧钩婊?    
    Returns:
        缁勫悎鎹熷け鍊?    """
    # CE 鎹熷け
    loss_ce = F.cross_entropy(logits, target, weight=weight, label_smoothing=label_smoothing)
    
    # QWK 鎹熷け
    loss_qwk = soft_qwk_loss(logits, target, num_classes)
    
    # 缁勫悎
    return (1.0 - qwk_weight) * loss_ce + qwk_weight * loss_qwk


def ordinal_regression_loss(
    logits: torch.Tensor,
    target: torch.Tensor,
    num_classes: int = 5,
    weight: Optional[torch.Tensor] = None,
) -> torch.Tensor:
    """
    绠€鍖栫殑鏈夊簭鍥炲綊鎹熷け
    
    鎯╃綒璺濈鐪熷疄鏍囩瓒婅繙鐨勯娴嬭秺閲?    
    Args:
        logits: (B, num_classes) 妯″瀷杈撳嚭
        target: (B,) 鐪熷疄鏍囩
        num_classes: 绫诲埆鏁伴噺
        weight: 绫诲埆鏉冮噸
    
    Returns:
        鎹熷け鍊?    """
    device = logits.device
    batch_size = logits.shape[0]
    
    # Soft prediction distribution.
    probs = F.softmax(logits, dim=1)  # (B, C)
    
    # Expected predicted grade.
    class_indices = torch.arange(num_classes, device=device).float()  # (C,)
    pred_expected = (probs * class_indices).sum(dim=1)  # (B,)
    
    # 鐪熷疄鏍囩
    target_float = target.float()  # (B,)
    
    # 骞虫柟璇樊 (绫讳技 MSE锛屼絾瀵逛簬鍒嗙被杈撳嚭)
    loss = (pred_expected - target_float) ** 2
    
    # 鍙€? 绫诲埆鏉冮噸
    if weight is not None:
        sample_weight = weight[target.clamp(0, num_classes - 1)]
        loss = loss * sample_weight
    
    return loss.mean()

def grade_regression_metrics(pred_scores: torch.Tensor, y: torch.Tensor) -> Dict[str, float]:
    """
    Metrics for o_O-style DR regression:
    - pred_scores: shape [N] or [N, 1], real-valued predictions in [0, 4] (not enforced)
    - y: int labels in {0, 1, 2, 3, 4}
    We apply fixed thresholds (0.5, 1.5, 2.5, 3.5) to map scores back to 5 discrete levels,
    then compute accuracy, macro-F1, and quadratic weighted kappa.
    """
    if pred_scores.ndim == 2 and pred_scores.shape[1] == 1:
        scores = pred_scores.view(-1).cpu().numpy()
    else:
        scores = pred_scores.detach().cpu().numpy().reshape(-1)
    y_np = y.detach().cpu().numpy().astype(int)

    # Map continuous scores to discrete 0..4 levels (as in o_O: thresholds at 0.5, 1.5, 2.5, 3.5)
    pred = np.zeros_like(scores, dtype=int)
    pred[scores >= 0.5] = 1
    pred[scores >= 1.5] = 2
    pred[scores >= 2.5] = 3
    pred[scores >= 3.5] = 4

    out: Dict[str, float] = {}
    out["acc"] = float((pred == y_np).mean())
    if f1_score is not None:
        out["macro_f1"] = float(f1_score(y_np, pred, average="macro"))
    else:
        out["macro_f1"] = macro_f1_numpy(y_np, pred, 5)
    if cohen_kappa_score is not None and len(np.unique(y_np)) > 1:
        out["qwk"] = float(cohen_kappa_score(y_np, pred, weights="quadratic"))
    return out

def extract_patient_id(path_str: str) -> str:
    stem = Path(path_str).stem
    # Common Kaggle-style: "<patient>_left", "<patient>_right", "<patient>_XXXX"
    m = re.match(r"^(\d+)_", stem)
    if m:
        return m.group(1)

    # Project naming convention: "<dataset>__<modality>__<key>__<hash>"
    # Examples:
    #   eyepacs__fundus__10003_left__99147a  -> 10003
    #   deepdrid__fundus__347_l1__583051     -> 347
    #   idrid__seg__IDRiD_01__ddc747         -> IDRiD_01
    parts = stem.split("__")
    if len(parts) >= 3:
        key = parts[2]
        m2 = re.match(r"^(.*)_(left|right|l\d*|r\d*)$", key, flags=re.IGNORECASE)
        if m2:
            return m2.group(1)
        return key

    return stem


def infer_eye_side(path_str: str) -> int:
    """Infer eye side: 1 if right, 0 otherwise."""
    s = Path(path_str).stem.lower()
    if "right" in s or s.endswith("_r") or "_r_" in s:
        return 1
    if "left" in s or s.endswith("_l") or "_l_" in s:
        return 0
    return 0


def split_rows_stratified(rows: List[Dict[str, str]], label_key_candidates: List[str], seed: int, test_size: float, val_size: float) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    keys = list(rows[0].keys())
    label_key = get_first_existing_key(keys, label_key_candidates)
    if label_key is None:
        raise ValueError(f"Label column not found. Found keys: {sorted(keys)}")

    test_size = float(test_size)
    val_size = float(val_size)
    if not (0.0 < test_size < 1.0) or not (0.0 < val_size < 1.0) or (test_size + val_size) >= 1.0:
        raise ValueError(f"Invalid split fractions: test_size={test_size} val_size={val_size} (must satisfy 0<.. and test+val<1).")

    # Use only labeled rows for splitting.
    y_all = np.array([int(float(r.get(label_key, -1))) for r in rows], dtype=np.int64)
    idx_all = np.arange(len(rows), dtype=np.int64)
    labeled_mask = y_all >= 0
    idx_lab = idx_all[labeled_mask]
    y_lab = y_all[labeled_mask]
    if int(idx_lab.size) < 10:
        raise ValueError("Not enough labeled samples for splitting.")

    # Group by patient (so left/right eyes stay in the same split).
    group_ids = np.array([extract_patient_id(rows[i]["image_path"]) for i in idx_lab], dtype=object)

    # Build group -> indices, and group-level label for stratification.
    by_group: Dict[str, List[int]] = {}
    by_group_y: Dict[str, List[int]] = {}
    for i_row, y_i, g in zip(idx_lab.tolist(), y_lab.tolist(), group_ids.tolist()):
        g = str(g)
        by_group.setdefault(g, []).append(int(i_row))
        by_group_y.setdefault(g, []).append(int(y_i))

    uniq_groups = np.array(list(by_group.keys()), dtype=object)
    g_y = np.array([int(max(by_group_y[g])) for g in uniq_groups.tolist()], dtype=np.int64)

    # Split at group-level with patient grouping + label stratification (when possible).
    cnt = np.bincount(g_y, minlength=int(g_y.max()) + 1 if g_y.size else 0)
    cnt_nz = cnt[cnt > 0]
    can_stratify = (StratifiedShuffleSplit is not None) and (cnt_nz.size >= 2) and (int(cnt_nz.min()) >= 2)

    groups_tr: Set[str]
    groups_va: Set[str]
    groups_te: Set[str]
    val_frac = val_size / (1.0 - test_size)

    if not can_stratify:
        if StratifiedShuffleSplit is None:
            why = "sklearn_missing"
        elif cnt_nz.size < 2:
            why = "single_label"
        else:
            why = "rare_label"
        print(
            f"[SPLIT][WARN] Falling back to patient-group split without label stratification ({why}). "
            f"patient_label_counts={cnt.tolist()}",
            flush=True,
        )
        rng = np.random.RandomState(seed)
        perm = rng.permutation(len(uniq_groups))
        n_test = max(1, int(round(len(perm) * test_size)))
        groups_te = set(str(g) for g in uniq_groups[perm[:n_test]].tolist())
        trval_groups = uniq_groups[perm[n_test:]]

        rng2 = np.random.RandomState(seed + 1)
        perm2 = rng2.permutation(len(trval_groups))
        n_val = max(1, int(round(len(trval_groups) * val_frac)))
        groups_va = set(str(g) for g in trval_groups[perm2[:n_val]].tolist())
        groups_tr = set(str(g) for g in trval_groups[perm2[n_val:]].tolist())
    else:
        if cnt_nz.size and int(cnt_nz.min()) < 5:
            print(f"[SPLIT][WARN] Some labels have few patients: counts={cnt.tolist()}. Stratified split may be a bit noisy.", flush=True)

        X_dummy = np.zeros((len(uniq_groups), 1), dtype=np.float32)
        # 1) test split
        try:
            sss1 = StratifiedShuffleSplit(n_splits=1, test_size=test_size, random_state=seed)
            g_trval_i, g_te_i = next(sss1.split(X_dummy, g_y))
        except Exception as e:
            print(f"[SPLIT][WARN] Stratified patient split failed (test): {e}. Falling back to group-only split.", flush=True)
            rng = np.random.RandomState(seed)
            perm = rng.permutation(len(uniq_groups))
            n_test = max(1, int(round(len(perm) * test_size)))
            groups_te = set(str(g) for g in uniq_groups[perm[:n_test]].tolist())
            trval_groups = uniq_groups[perm[n_test:]]
            rng2 = np.random.RandomState(seed + 1)
            perm2 = rng2.permutation(len(trval_groups))
            n_val = max(1, int(round(len(trval_groups) * val_frac)))
            groups_va = set(str(g) for g in trval_groups[perm2[:n_val]].tolist())
            groups_tr = set(str(g) for g in trval_groups[perm2[n_val:]].tolist())
        else:
            groups_trval = uniq_groups[g_trval_i]
            y_trval = g_y[g_trval_i]
            groups_te = set(str(g) for g in uniq_groups[g_te_i].tolist())

            # 2) val split from remaining
            try:
                sss2 = StratifiedShuffleSplit(n_splits=1, test_size=val_frac, random_state=seed + 1)
                X2 = np.zeros((len(groups_trval), 1), dtype=np.float32)
                g_tr_i, g_va_i = next(sss2.split(X2, y_trval))
            except Exception as e:
                print(f"[SPLIT][WARN] Stratified patient split failed (val): {e}. Falling back to group-only split.", flush=True)
                rng2 = np.random.RandomState(seed + 1)
                perm2 = rng2.permutation(len(groups_trval))
                n_val = max(1, int(round(len(groups_trval) * val_frac)))
                groups_va = set(str(g) for g in groups_trval[perm2[:n_val]].tolist())
                groups_tr = set(str(g) for g in groups_trval[perm2[n_val:]].tolist())
            else:
                groups_tr = set(str(g) for g in groups_trval[g_tr_i].tolist())
                groups_va = set(str(g) for g in groups_trval[g_va_i].tolist())

    # Materialize row splits.
    tr_idx: List[int] = []
    va_idx: List[int] = []
    te_idx: List[int] = []
    for g, indices in by_group.items():
        if g in groups_tr:
            tr_idx.extend(indices)
        elif g in groups_va:
            va_idx.extend(indices)
        elif g in groups_te:
            te_idx.extend(indices)

    train_rows = [rows[i] for i in tr_idx]
    val_rows = [rows[i] for i in va_idx]
    test_rows = [rows[i] for i in te_idx]
    return train_rows, val_rows, test_rows


# -----------------------------
# Training config
# -----------------------------
@dataclass
# -----------------------------
# Training config
# -----------------------------
@dataclass
class TrainCfg:
    task: str
    project_root: Path
    run_dir: Path
    backbone: str
    roundtune_ckpt: Optional[Path]
    init_encoder_ckpt: Optional[Path]
    init_quality_ckpt: Optional[Path]
    timm_pretrained_file: Optional[Path]
    img_size: int
    batch_size: int
    epochs: int
    lr: float
    weight_decay: float
    dropout: float
    num_workers: int
    amp: bool
    seed: int
    device: str
    tensorboard: bool
    grad_clip: float
    accumulate_steps: int
    limit: int
    log_every_steps: int
    tune_mode: str
    unfreeze_last_n: int
    lora_r: int
    lora_alpha: float
    lora_dropout: float
    lora_targets: str
    lora_lr_mult: float
    lora_wd_mult: float
    head_lr_mult: float
    num_classes_quality: int
    qual_use_weighted_sampler: bool
    qual_use_class_weights: bool
    qual_gradable_index: int
    qual_collapse_to_binary: bool
    grade_loss: str
    grade_focal_gamma: float
    qwk_loss_weight: float
    edl_evidence: str
    edl_kl_weight: float
    edl_anneal_epochs: int
    grade_blend: bool
    grade_blend_tta: int
    grade_blend_epochs: int
    grade_aug: str
    grade_sampler: str
    sampler_stage1: str
    sampler_stage2: str
    sampler_stage2_start_frac: float
    train_use_ce_weight: bool
    val_use_ce_weight: bool
    early_stop_min_epoch: int
    patience: int
    min_delta: float


def build_optimizer(model: nn.Module, cfg: TrainCfg) -> torch.optim.Optimizer:
    """Build optimizer for image-only quality/grade models.

    Parameter groups:
      A) backbone_non_lora: trainable backbone params excluding LoRA
      B) lora_params: LoRA adapter params
      C) head_params: non-backbone classifier/regression heads
    """
    lr = float(getattr(cfg, "lr", 1e-4))
    wd = float(getattr(cfg, "weight_decay", 0.0))
    lora_lr_mult = float(getattr(cfg, "lora_lr_mult", 0.1))
    lora_wd_mult = float(getattr(cfg, "lora_wd_mult", 1.0))
    head_lr_mult = float(getattr(cfg, "head_lr_mult", 1.0))

    backbone_lr = lr * 0.1
    lora_lr = lr * lora_lr_mult
    head_lr = lr * head_lr_mult

    backbone_non_lora: List[nn.Parameter] = []
    lora_params: List[nn.Parameter] = []
    head_params: List[nn.Parameter] = []
    seen: Set[int] = set()

    for name, param in model.named_parameters():
        if not param.requires_grad:
            continue
        pid = id(param)
        if pid in seen:
            continue
        seen.add(pid)

        lname = name.lower()
        if "lora" in lname:
            lora_params.append(param)
        elif name.startswith("backbone."):
            backbone_non_lora.append(param)
        else:
            head_params.append(param)

    def _numel(params: List[nn.Parameter]) -> int:
        return int(sum(int(p.numel()) for p in params))

    print(
        "[OPT] param_groups:",
        f"backbone_non_lora n={len(backbone_non_lora)} numel={_numel(backbone_non_lora)} lr={backbone_lr:.2e} wd={wd:g};",
        f"lora n={len(lora_params)} numel={_numel(lora_params)} lr={lora_lr:.2e} wd={wd * lora_wd_mult:g};",
        f"head n={len(head_params)} numel={_numel(head_params)} lr={head_lr:.2e} wd={wd:g}",
        flush=True,
    )

    param_groups: List[Dict[str, Any]] = []
    if backbone_non_lora:
        param_groups.append({"params": backbone_non_lora, "lr": backbone_lr, "weight_decay": wd})
    if lora_params:
        param_groups.append({"params": lora_params, "lr": lora_lr, "weight_decay": wd * lora_wd_mult})
    if head_params:
        param_groups.append({"params": head_params, "lr": head_lr, "weight_decay": wd})
    if not param_groups:
        raise ValueError("[OPT] No trainable parameters found. Check tune_mode/requires_grad settings.")

    return torch.optim.AdamW(param_groups)


def _count_params(model: nn.Module) -> Tuple[int, int]:
    total = 0
    trainable = 0
    for p in model.parameters():
        n = int(p.numel())
        total += n
        if p.requires_grad:
            trainable += n
    return total, trainable


def is_grade_transfer_run(*, has_encoder_init: bool, use_timm_pretrained: bool) -> bool:
    """Return True when grade training starts from any pretrained image encoder."""
    return bool(has_encoder_init or use_timm_pretrained)


def apply_lora_to_mae_vit_backbone(
    backbone: "SharedBackbone",
    *,
    r: int,
    alpha: float = 1.0,
    dropout: float = 0.0,
    targets: Optional[Set[str]] = None,
) -> int:
    """
    Apply LoRA to MAE-ViT encoder blocks.
    targets: {"qkv","proj"} (default). Returns number of wrapped Linear modules.
    """
    tset = {t.strip().lower() for t in (targets or {"qkv", "proj"}) if t}
    bb = getattr(backbone, "backbone", None)
    blocks = getattr(bb, "blocks", None)
    if blocks is None:
        return 0
    n_wrapped = 0
    for blk in blocks:
        attn = getattr(blk, "attn", None)
        if attn is None:
            continue
        if "qkv" in tset:
            qkv = getattr(attn, "qkv", None)
            if isinstance(qkv, LoRALinear):
                n_wrapped += 1
            elif isinstance(qkv, nn.Linear):
                attn.qkv = LoRALinear(qkv, r=int(r), alpha=float(alpha), dropout=float(dropout))
                n_wrapped += 1
        if "proj" in tset:
            proj = getattr(attn, "proj", None)
            if isinstance(proj, LoRALinear):
                n_wrapped += 1
            elif isinstance(proj, nn.Linear):
                attn.proj = LoRALinear(proj, r=int(r), alpha=float(alpha), dropout=float(dropout))
                n_wrapped += 1
    return n_wrapped


def apply_finetune_strategy(model: nn.Module, cfg: TrainCfg) -> Dict[str, Any]:
    """
    Configure trainable parameters for transfer learning experiments:
      - full: train everything
      - freeze: freeze encoder/backbone
      - ln: freeze encoder except LayerNorm params
      - last_n: freeze encoder except last N blocks (+final norm)
      - lora: freeze encoder weights; train LoRA(qkv,proj) + heads/decoders
    """
    mode = str(getattr(cfg, "tune_mode", "full")).lower().strip()
    info: Dict[str, Any] = {"tune_mode": mode}

    if mode in ("", "full", "none"):
        total, trainable = _count_params(model)
        info.update({"trainable_params": trainable, "total_params": total})
        print(f"[FT] tune=full trainable={trainable/1e6:.2f}M total={total/1e6:.2f}M", flush=True)
        return info

    # 1) Freeze encoder
    for n, p in model.named_parameters():
        if n.startswith("backbone."):
            p.requires_grad = False

    # 2) Optional unfreezing strategies
    if mode == "ln":
        for n, p in model.named_parameters():
            if not n.startswith("backbone."):
                continue
            # MAE-ViT uses LayerNorm named norm/norm1/norm2; keep this name-based for minimal coupling.
            if ".norm" in n and (n.endswith(".weight") or n.endswith(".bias")):
                p.requires_grad = True

    elif mode == "last_n":
        n_last = int(getattr(cfg, "unfreeze_last_n", 0))
        if n_last <= 0:
            raise ValueError("--unfreeze_last_n must be >0 for --tune_mode last_n")
        bb = getattr(getattr(model, "backbone", None), "backbone", None)
        blocks = getattr(bb, "blocks", None)
        if blocks is None:
            raise ValueError("[FT] last_n requires a ViT-like backbone with .blocks")
        depth = len(blocks)
        start = max(0, depth - n_last)
        for i in range(start, depth):
            prefix = f"backbone.backbone.blocks.{i}."
            for n, p in model.named_parameters():
                if n.startswith(prefix):
                    p.requires_grad = True
        # Always allow final norm to adapt a bit.
        for n, p in model.named_parameters():
            if n.startswith("backbone.backbone.norm."):
                p.requires_grad = True
        info.update({"unfreeze_last_n": n_last, "depth": depth})

    elif mode == "lora":
        r = int(getattr(cfg, "lora_r", 0))
        if r <= 0:
            raise ValueError("--lora_r must be >0 for --tune_mode lora")
        alpha = float(getattr(cfg, "lora_alpha", 1.0))
        dropout = float(getattr(cfg, "lora_dropout", 0.0))
        targets_s = str(getattr(cfg, "lora_targets", "qkv,proj"))
        targets = {t.strip().lower() for t in targets_s.split(",") if t.strip()}
        n_wrapped = apply_lora_to_mae_vit_backbone(getattr(model, "backbone"), r=r, alpha=alpha, dropout=dropout, targets=targets)
        # Ensure LoRA params remain trainable even though they live under backbone.*
        for n, p in model.named_parameters():
            if n.startswith("backbone.") and ("lora_" in n):
                p.requires_grad = True
        info.update({"lora_r": r, "lora_alpha": alpha, "lora_dropout": dropout, "lora_targets": sorted(list(targets)), "lora_wrapped": int(n_wrapped)})
        print(f"[FT] LoRA applied: wrapped={n_wrapped} targets={sorted(list(targets))} r={r} alpha={alpha} dropout={dropout}", flush=True)

    elif mode == "freeze":
        pass
    else:
        raise ValueError(f"Unsupported --tune_mode '{mode}'. Choose from: full, freeze, ln, last_n, lora")

    total, trainable = _count_params(model)
    info.update({"trainable_params": trainable, "total_params": total})
    print(f"[FT] tune={mode} trainable={trainable/1e6:.2f}M total={total/1e6:.2f}M", flush=True)
    return info


def save_encoder_only(backbone: SharedBackbone, path: Path) -> None:
    torch.save(backbone.state_dict(), str(path))


def choose_backbone_name(cfg: TrainCfg) -> str:
    """Prefer backbone inferred from provided checkpoints to avoid mismatched loading."""
    # 1) init_encoder_ckpt has highest priority (e.g., quality encoder you want to reuse)
    if cfg.init_encoder_ckpt is not None and cfg.init_encoder_ckpt.exists():
        st = load_checkpoint_state(cfg.init_encoder_ckpt)
        inferred = infer_backbone_from_state_dict_keys(list(st.keys()), st)
        if inferred and inferred != cfg.backbone:
            print(f"[INIT] overriding backbone '{cfg.backbone}' -> '{inferred}' based on init_encoder_ckpt")
            return inferred
    # 2) fallback to RoundTune ckpt inference
    if cfg.roundtune_ckpt is not None and cfg.roundtune_ckpt.exists():
        state, inferred = load_roundtune_encoder_weights(cfg.roundtune_ckpt)
        inferred = inferred or infer_backbone_from_state_dict_keys(list(state.keys()), state)
        if inferred and inferred != cfg.backbone:
            print(f"[INIT] overriding backbone '{cfg.backbone}' -> '{inferred}' based on RoundTune ckpt")
            return inferred
    return cfg.backbone


def init_backbone(backbone: SharedBackbone, cfg: TrainCfg) -> Dict[str, Any]:
    init_info: Dict[str, Any] = {"init": "scratch"}
    is_mae = cfg.backbone.lower().startswith("mae_vit")

    def _state_has_lora_wrapped_keys(st: Dict[str, torch.Tensor]) -> bool:
        # Our LoRALinear wrapper saves weights as:
        #   *.base.weight / *.base.bias / *.lora_A / *.lora_B / *.lora_alpha
        return any((".base." in k) or (".lora_" in k) for k in st.keys())

    def _maybe_prepare_mae_for_lora_ckpt(st: Dict[str, torch.Tensor], src: str) -> Dict[str, torch.Tensor]:
        """
        If the checkpoint was saved from a LoRA-wrapped MAE backbone:
          - tune_mode=lora: apply LoRA wrappers BEFORE loading so keys match.
          - otherwise: strip LoRA adapters and map *.base.* -> * for best-effort loading.
        """
        if not _state_has_lora_wrapped_keys(st):
            return st

        mode = str(getattr(cfg, "tune_mode", "full")).lower().strip()
        if mode == "lora":
            r = int(getattr(cfg, "lora_r", 0) or 0)
            alpha = float(getattr(cfg, "lora_alpha", 1.0))
            dropout = float(getattr(cfg, "lora_dropout", 0.0))
            targets_s = str(getattr(cfg, "lora_targets", "qkv,proj"))
            targets = {t.strip().lower() for t in targets_s.split(",") if t.strip()}
            if r > 0:
                n_wrapped = apply_lora_to_mae_vit_backbone(
                    backbone,
                    r=r,
                    alpha=alpha,
                    dropout=dropout,
                    targets=targets,
                )
                print(
                    f"[INIT] detected LoRA-wrapped MAE encoder in {src}; "
                    f"applied LoRA wrappers before loading (wrapped={n_wrapped})",
                    flush=True,
                )
            return st

        # Non-LoRA tuning: load the base weights (drop adapters) to reduce missing/unexpected.
        print(
            f"[INIT] detected LoRA-wrapped MAE encoder in {src} but tune_mode='{mode}'; "
            "stripping LoRA adapters and mapping '*.base.*' -> '*' for loading",
            flush=True,
        )
        out: Dict[str, torch.Tensor] = {}
        for k, v in st.items():
            if ".lora_" in k:
                continue
            if ".base." in k:
                out[k.replace(".base.", ".")] = v
            else:
                out[k] = v
        return out

    if cfg.init_encoder_ckpt is not None and cfg.init_encoder_ckpt.exists():
        st = load_checkpoint_state(cfg.init_encoder_ckpt)
        if not is_mae:
            st = _remap_convnext_underscored(st)
            st = _maybe_prefix_backbone(st)
            info = safe_load(backbone, st, strict=False)
        else:
            st = _maybe_prepare_mae_for_lora_ckpt(st, src="init_encoder_ckpt")
            info, st = _try_load_variants(backbone, st)
            bad = set(info.get("shape_mismatch", []))
            st_load = {k: v for k, v in st.items() if k not in bad}
            st_load, dropped_alpha = drop_lora_alpha_for_mismatched_adapters(st_load, info.get("shape_mismatch", []))
            load_info = safe_load(backbone, st_load, strict=False)
            info = {**info, **load_info}
            if dropped_alpha:
                info["dropped_lora_alpha_due_to_adapter_shape_mismatch"] = dropped_alpha
        init_info = {"init": "init_encoder_ckpt", "path": str(cfg.init_encoder_ckpt), **info}
        print(f"[INIT] encoder from init_encoder_ckpt (variant={init_info.get('variant','')}) "
              f"missing={len(info['missing'])} unexpected={len(info['unexpected'])}")
    elif cfg.roundtune_ckpt is not None and cfg.roundtune_ckpt.exists():
        st, inferred = load_roundtune_encoder_weights(cfg.roundtune_ckpt)
        if not is_mae:
            st = _remap_convnext_underscored(st)
            st = _maybe_prefix_backbone(st)
            info = safe_load(backbone, st, strict=False)
        else:
            st = _maybe_prepare_mae_for_lora_ckpt(st, src="roundtune_ckpt")
            info, st = _try_load_variants(backbone, st)
            bad = set(info.get("shape_mismatch", []))
            st_load = {k: v for k, v in st.items() if k not in bad}
            st_load, dropped_alpha = drop_lora_alpha_for_mismatched_adapters(st_load, info.get("shape_mismatch", []))
            load_info = safe_load(backbone, st_load, strict=False)
            info = {**info, **load_info}
            if dropped_alpha:
                info["dropped_lora_alpha_due_to_adapter_shape_mismatch"] = dropped_alpha
        init_info = {"init": "roundtune", "path": str(cfg.roundtune_ckpt), "inferred_backbone": inferred, **info}
        print(f"[INIT] encoder from RoundTune (inferred={inferred}, variant={info.get('variant','')}) "
              f"missing={len(info['missing'])} unexpected={len(info['unexpected'])}")
    else:
        print("[INIT] WARNING: no encoder init checkpoint provided. Training from scratch.")
    return init_info


# -----------------------------
# Train: Quality


# -----------------------------
def train_qual(cfg: TrainCfg, qual_manifest: Path) -> None:
    rows = read_csv_rows(qual_manifest)
    if not rows:
        raise ValueError(f"[QUAL] Empty manifest: {qual_manifest}")

    # -----------------------------
    # 1) Normalize / collapse labels into a safe "label" column (0..K-1)
    # -----------------------------
    keys = list(rows[0].keys())
    raw_key = get_first_existing_key(keys, ["label", "quality", "overall", "gradability", "label_quality", "q"])
    if raw_key is None:
        raise ValueError(f"[QUAL] No label column found. Keys={sorted(keys)}")

    y_raw: List[int] = []
    for r in rows:
        v = r.get(raw_key, "")
        if v == "" or v is None:
            y_raw.append(-1)
        else:
            y_raw.append(int(float(v)))

    uniq = sorted({y for y in y_raw if y >= 0})
    if len(uniq) == 0:
        raise ValueError("[QUAL] No labeled samples found in manifest (all labels empty).")

    collapse = bool(cfg.qual_collapse_to_binary)

    # Auto-collapse safeguard: user requested 2-class but manifest looks multi-class
    if int(cfg.num_classes_quality) == 2 and (not collapse) and len(uniq) > 2:
        print(f"[QUAL] WARNING: detected {len(uniq)} unique labels {uniq} but num_classes_quality=2. "
              f"Auto-enable --qual_collapse_to_binary for this run.")
        collapse = True

    label_map: Dict[int, int] = {}

    if int(cfg.num_classes_quality) == 2 and collapse:
        # Common 3-class conventions:
        #   A) 0=Good, 1=Usable, 2=Reject  -> gradable={0,1} -> 1, reject -> 0
        #   B) 1=Good, 2=Usable, 3=Reject  -> gradable={1,2} -> 1, reject -> 0
        # If labels already binary {0,1}, keep unchanged.
        if set(uniq).issubset({0, 1}):
            def _to_bin(y: int) -> int:
                return y
        else:
            # choose rule by checking whether label '0' exists
            if 0 in uniq:
                gradable_set = {0, 1}
            else:
                gradable_set = {1, 2}
            def _to_bin(y: int) -> int:
                return 1 if y in gradable_set else 0

        y_proc = [(_to_bin(y) if y >= 0 else -1) for y in y_raw]
        proc_uniq = sorted({y for y in y_proc if y >= 0})
        if proc_uniq != [0, 1] and proc_uniq != [0] and proc_uniq != [1]:
            raise ValueError(f"[QUAL] Binary collapse produced unexpected labels: {proc_uniq}")
        num_classes = 2

    else:
        # No collapse: map raw unique labels to consecutive 0..K-1
        num_classes = int(cfg.num_classes_quality)
        # If labels already within range, mapping is identity.
        if min(uniq) >= 0 and max(uniq) < num_classes:
            label_map = {u: u for u in uniq}
        else:
            # Typical: labels are 1..K, map to 0..K-1
            label_map = {u: i for i, u in enumerate(uniq)}

        y_proc = [(label_map[y] if y >= 0 else -1) for y in y_raw]
        proc_uniq = sorted({y for y in y_proc if y >= 0})
        if len(proc_uniq) > num_classes:
            raise ValueError(f"[QUAL] After mapping, got {len(proc_uniq)} labels {proc_uniq} "
                             f"but num_classes_quality={num_classes}. Please set --num_classes_quality correctly.")
        if max(proc_uniq) >= num_classes:
            raise ValueError(f"[QUAL] Mapped labels out of range: {proc_uniq} with num_classes={num_classes}")

    # Write processed labels into a new column "label" (preferred by dataset)
    rows2: List[Dict[str, str]] = []
    for r, y in zip(rows, y_proc):
        rr = dict(r)
        rr["label"] = "" if y < 0 else str(int(y))
        rows2.append(rr)

    print(f"[QUAL] label_col='{raw_key}' -> normalized to 'label' with uniq={sorted({y for y in y_proc if y >= 0})} "
          f"(num_classes={num_classes}) collapse={collapse}")

    # -----------------------------
    # 2) Split (patient-aware if possible)
    # -----------------------------
    tr_rows, va_rows, te_rows = split_rows_stratified(
        rows2,
        ["label"],  # use normalized label
        cfg.seed,
        test_size=0.1,
        val_size=0.2,
    )

    ds_tr = QualityDataset(tr_rows, cfg.project_root, cfg.img_size, True, collapse_to_binary=False)
    ds_va = QualityDataset(va_rows, cfg.project_root, cfg.img_size, False, collapse_to_binary=False)
    ds_te = QualityDataset(te_rows, cfg.project_root, cfg.img_size, False, collapse_to_binary=False)

    # -----------------------------
    # 3) Class balancing (loss weights + sampler)
    # -----------------------------
    class_weights_tensor: Optional[torch.Tensor] = None
    sampler = None

    y_list: List[int] = []
    for r in tr_rows:
        v = r.get("label", "")
        y = -1 if (v == "" or v is None) else int(float(v))
        y_list.append(y)

    y_arr = np.array([y for y in y_list if y >= 0], dtype=np.int64)
    if y_arr.size > 0:
        counts = np.bincount(y_arr, minlength=num_classes)
        print(f"[QUAL] class counts (train) = {counts.tolist()}  num_classes={num_classes}")

        if cfg.qual_use_class_weights:
            cw = counts.sum() / np.maximum(counts, 1)
            cw = cw / cw.mean()  # normalize
            class_weights_tensor = torch.tensor(cw, dtype=torch.float32, device=cfg.device)
            print(f"[QUAL] class weights (loss) = {[round(float(x), 3) for x in cw.tolist()]}")

        if cfg.qual_use_weighted_sampler:
            try:
                from torch.utils.data import WeightedRandomSampler  # type: ignore
                inv = counts.sum() / np.maximum(counts, 1)
                w = [float(inv[y]) if y >= 0 else 0.0 for y in y_list]
                w_t = torch.tensor(w, dtype=torch.double)
                sampler = WeightedRandomSampler(weights=w_t, num_samples=len(w_t), replacement=True)
                print("[QUAL] WeightedRandomSampler enabled.")
            except Exception as e:
                print(f"[QUAL] WeightedRandomSampler unavailable: {e}")
                sampler = None
    else:
        print("[QUAL] WARNING: no labeled samples found for quality training (after split).")

    dl_tr = DataLoader(
        ds_tr,
        batch_size=cfg.batch_size,
        shuffle=(sampler is None),
        sampler=sampler,
        num_workers=cfg.num_workers,
        pin_memory=True,
        drop_last=True,
        persistent_workers=(cfg.num_workers > 0),
    )
    dl_va = DataLoader(
        ds_va,
        batch_size=cfg.batch_size,
        shuffle=False,
        num_workers=cfg.num_workers,
        pin_memory=True,
        drop_last=False,
        persistent_workers=(cfg.num_workers > 0),
    )
    dl_te = DataLoader(
        ds_te,
        batch_size=cfg.batch_size,
        shuffle=False,
        num_workers=cfg.num_workers,
        pin_memory=True,
        drop_last=False,
        persistent_workers=(cfg.num_workers > 0),
    )

    backbone_name = choose_backbone_name(cfg)
    use_pretrained = not ((cfg.roundtune_ckpt is not None and cfg.roundtune_ckpt.exists()) or (cfg.init_encoder_ckpt is not None and cfg.init_encoder_ckpt.exists()))
    cfg.backbone = backbone_name
    backbone = SharedBackbone(backbone_name)
    init_info = init_backbone(backbone, cfg)

    model = QualityModel(backbone, num_classes_quality=num_classes, dropout=0.2).to(cfg.device)
    init_info = {**init_info, "finetune": apply_finetune_strategy(model, cfg)}
    opt = build_optimizer(model, cfg)
    scaler = amp_grad_scaler(cfg.device, cfg.amp)
    writer = build_writer(cfg.run_dir, cfg.tensorboard)

    logs_dir = cfg.run_dir / "logs"
    hist_csv = logs_dir / "history.csv"
    hist_jsonl = logs_dir / "history.jsonl"
    summary_json = logs_dir / "summary.json"
    fields = ["task", "epoch", "split", "loss", "acc", "macro_f1", "time_sec"]
    init_history_file(hist_csv, fields)
    save_json(summary_json, {"task": "qual", "init": init_info, "config": cfg.__dict__})

    best_score = -1.0
    best_epoch = 0

    patience = max(0, int(cfg.patience))
    min_delta = float(cfg.min_delta)
    bad_epochs = 0

    for epoch in range(1, cfg.epochs + 1):
        model.train()
        t0 = time.time()
        opt.zero_grad(set_to_none=True)

        losses = []
        loss_sum = 0.0
        steps_per_epoch = len(dl_tr)
        log_every = max(0, int(cfg.log_every_steps))
        train_correct = 0
        train_total = 0
        train_y_list: List[int] = []
        train_pred_list: List[int] = []
        for step, batch in enumerate(dl_tr, 1):
            x = batch["image"].to(cfg.device, non_blocking=True)
            y = batch["label"].to(cfg.device, non_blocking=True)
            has_label = batch["has_label"].to(cfg.device, non_blocking=True).float()

            with amp_autocast(cfg.device, cfg.amp):
                logits, _ = model(x)
                # Label smoothing=0.05 for quality (smaller than grade since labels are cleaner)
                loss_raw = F.cross_entropy(
                    logits,
                    y.clamp_min(0),
                    reduction="none",
                    weight=class_weights_tensor if class_weights_tensor is not None else None,
                    label_smoothing=0.05,
                )
                loss = (loss_raw * has_label).sum() / (has_label.sum().clamp(min=1.0))
                loss = loss / cfg.accumulate_steps

            mask = has_label > 0
            if mask.any():
                pred = torch.argmax(logits.detach(), dim=1)
                train_correct += int((pred[mask] == y[mask]).sum().item())
                train_total += int(mask.sum().item())
                train_y_list.extend(y[mask].detach().cpu().tolist())
                train_pred_list.extend(pred[mask].detach().cpu().tolist())

            scaler.scale(loss).backward()
            if step % cfg.accumulate_steps == 0:
                if cfg.grad_clip > 0:
                    scaler.unscale_(opt)
                    nn.utils.clip_grad_norm_(model.parameters(), cfg.grad_clip)
                scaler.step(opt)
                scaler.update()
                opt.zero_grad(set_to_none=True)

            loss_item = float(loss.item()) * cfg.accumulate_steps
            losses.append(loss_item)
            loss_sum += loss_item
            if log_every > 0 and (step == 1 or step % log_every == 0 or step == steps_per_epoch):
                elapsed = time.time() - t0
                eta_sec = (elapsed / max(1, step)) * (steps_per_epoch - step)
                lr_now = float(opt.param_groups[-1].get("lr", cfg.lr)) if opt.param_groups else float(cfg.lr)
                label_frac = float(has_label.mean().detach().cpu().item())
                print(
                    f"[QUAL][Epoch {epoch}/{cfg.epochs}][{step}/{steps_per_epoch}] "
                    f"loss={loss_item:.4f} avg={loss_sum / step:.4f} lab={label_frac:.2f} lr={lr_now:.2e} eta={eta_sec/60:.1f}m",
                    flush=True,
                )

        train_loss = float(np.mean(losses)) if losses else 0.0
        train_acc = float(train_correct / max(1, train_total))
        train_f1 = np.nan
        if train_total > 0:
            try:
                y_np = np.array(train_y_list, dtype=np.int64)
                p_np = np.array(train_pred_list, dtype=np.int64)
                if f1_score is not None:
                    train_f1 = float(f1_score(y_np, p_np, average="macro"))
                else:
                    train_f1 = macro_f1_numpy(y_np, p_np, num_classes)
            except Exception:
                train_f1 = np.nan

        model.eval()
        all_logits, all_y = [], []
        val_losses: List[float] = []
        with torch.no_grad():
            for batch in dl_va:
                x = batch["image"].to(cfg.device, non_blocking=True)
                y = batch["label"].to(cfg.device, non_blocking=True)
                has_label = batch["has_label"].to(cfg.device, non_blocking=True)
                mask = has_label.bool()
                with amp_autocast(cfg.device, cfg.amp):
                    logits, _ = model(x)
                    loss_raw = F.cross_entropy(
                        logits,
                        y.clamp_min(0),
                        reduction="none",
                        weight=class_weights_tensor if class_weights_tensor is not None else None,
                    )
                has = has_label.float()
                if mask.any():
                    all_logits.append(logits[mask].detach().float().cpu())
                    all_y.append(y[mask].detach().cpu())
                    denom = has.sum().clamp(min=1.0)
                    val_losses.append(float(((loss_raw * has).sum() / denom).item()))

        if all_logits:
            logits_cat = torch.cat(all_logits, dim=0)
            y_cat = torch.cat(all_y, dim=0)
            mets = cls_metrics(logits_cat, y_cat)
            val_acc = float(mets.get("acc", 0.0))
            val_f1 = float(mets.get("macro_f1", np.nan))
        else:
            val_acc, val_f1 = 0.0, np.nan
        val_loss = float(np.mean(val_losses)) if val_losses else 0.0

        dt = time.time() - t0
        print(f"[QUAL][Epoch {epoch}/{cfg.epochs}] train_loss={train_loss:.4f} val_acc={val_acc:.4f} val_f1={val_f1:.4f} time={dt:.1f}s", flush=True)

        append_history_row(hist_csv, {"task": "qual", "epoch": epoch, "split": "train", "loss": train_loss, "acc": train_acc, "macro_f1": train_f1, "time_sec": dt}, fields)
        append_history_row(hist_csv, {"task": "qual", "epoch": epoch, "split": "val", "loss": val_loss, "acc": val_acc, "macro_f1": val_f1, "time_sec": ""}, fields)
        append_jsonl(hist_jsonl, {"task": "qual", "epoch": epoch, "train_loss": train_loss, "train_acc": train_acc, "train_macro_f1": train_f1, "val_loss": val_loss, "val_acc": val_acc, "val_macro_f1": val_f1, "time_sec": dt})

        if writer is not None:
            writer.add_scalar("qual/train_loss", train_loss, epoch)
            writer.add_scalar("qual/train_acc", train_acc, epoch)
            if not np.isnan(train_f1):
                writer.add_scalar("qual/train_macro_f1", train_f1, epoch)
            writer.add_scalar("qual/val_loss", val_loss, epoch)
            writer.add_scalar("qual/val_acc", val_acc, epoch)
            if not np.isnan(val_f1):
                writer.add_scalar("qual/val_macro_f1", val_f1, epoch)

        score = val_acc if np.isnan(val_f1) else (0.7 * val_acc + 0.3 * val_f1)
        if score > (best_score + min_delta):
            best_score = score
            best_epoch = epoch
            torch.save({"model": model.state_dict(), "epoch": epoch, "val_score": best_score}, str(cfg.run_dir / "best_quality.pth"))
            save_encoder_only(model.backbone, cfg.run_dir / "best_encoder.pth")
            print(f"[QUAL] [OK] Saved best: val_score={best_score:.4f}")
            bad_epochs = 0
        else:
            bad_epochs += 1
            if patience > 0 and bad_epochs >= patience:
                print(f"[QUAL][EarlyStop] no improvement for {patience} epochs (min_delta={min_delta}). Stop at epoch {epoch}.")
                break

    if writer is not None:
        writer.close()

    # -----------------------------
    # 4) Internal test
    # -----------------------------
    model.eval()
    all_logits, all_y = [], []
    with torch.no_grad():
        for batch in dl_te:
            x = batch["image"].to(cfg.device, non_blocking=True)
            y = batch["label"].to(cfg.device, non_blocking=True)
            has_label = batch["has_label"].to(cfg.device, non_blocking=True)
            mask = has_label.bool()
            logits, _ = model(x)
            if mask.any():
                all_logits.append(logits[mask].detach().cpu())
                all_y.append(y[mask].detach().cpu())

    te_acc, te_f1 = 0.0, np.nan
    if all_logits:
        logits_cat = torch.cat(all_logits, dim=0)
        y_cat = torch.cat(all_y, dim=0)
        mets = cls_metrics(logits_cat, y_cat)
        te_acc = float(mets.get("acc", 0.0))
        te_f1 = float(mets.get("macro_f1", np.nan))

    print(f"[QUAL][Internal Test] acc={te_acc:.4f} macro_f1={te_f1:.4f}")

    extra = {"TestAcc": te_acc}
    if not np.isnan(te_f1):
        extra["TestMacroF1"] = float(te_f1)

    # Make table dataset_name reflect the actual manifest used (avoid misleading "DeepDRiD" when training on EyeQ).
    qm = str(qual_manifest.name).lower()
    if "eyeq" in qm:
        dataset_name = "EyeQ (quality labels on EyePACS images)"
    elif "clean_qual" in qm:
        dataset_name = "EyeQ-clean (Master_Train only)"
    elif "deepdrid" in qm:
        dataset_name = "DeepDRiD regular_fundus_images"
    else:
        dataset_name = f"qual_manifest={qual_manifest.name}"

    finalize_table(
        run_dir=cfg.run_dir,
        task_name="Image quality / gradability",
        dataset_name=dataset_name,
        cfg=cfg.__dict__,
        best_epoch=best_epoch,
        best_metric="ValScore",
        best_value=best_score,
        extra=extra,
    )

    save_json(
        summary_json,
        {
            "task": "qual",
            "init": init_info,
            "config": cfg.__dict__,
            "label_processing": {
                "raw_label_col": raw_key,
                "unique_raw": uniq,
                "collapse_to_binary": collapse,
                "label_map": label_map,
                "unique_processed": sorted({y for y in y_proc if y >= 0}),
                "num_classes_used": num_classes,
            },
            "best_epoch": best_epoch,
            "best_val_score": best_score,
            "internal_test": {"acc": te_acc, "macro_f1": te_f1},
        },
    )
    print(f"[QUAL] Done. Outputs in: {cfg.run_dir}")

# -----------------------------
# Train: Image-only grading
# -----------------------------
def train_grade(cfg: TrainCfg, grade_manifest: Path) -> None:
    rows = read_csv_rows(grade_manifest)
    # Prefer manifest-provided patient-level split (train/val_train/calib/test) when available.
    calib_rows: List[Dict[str, str]] = []
    if rows and any("split" in r for r in rows):
        tr_rows: List[Dict[str, str]] = []
        va_rows: List[Dict[str, str]] = []
        te_rows: List[Dict[str, str]] = []
        unknown = 0
        for r in rows:
            s = str(r.get("split", "") or "").strip().lower()
            if s in ("train", "tr", "training"):
                tr_rows.append(r)
            elif s in ("val_train", "val", "valid", "validation", "valtrain"):
                va_rows.append(r)
            elif s in ("calib", "calibration"):
                calib_rows.append(r)
            elif s in ("test", "te"):
                te_rows.append(r)
            else:
                # Fallback: keep old behavior (treat unknown/empty as train) but warn.
                tr_rows.append(r)
                unknown += 1
        print(
            f"[GRADE] manifest split: train={len(tr_rows)} val_train={len(va_rows)} calib={len(calib_rows)} test={len(te_rows)}",
            flush=True,
        )
        if unknown > 0:
            print(f"[GRADE] WARNING: {unknown} rows have unknown split; treated as train.", flush=True)
        if not tr_rows or not va_rows:
            raise ValueError("[GRADE] grade_manifest has 'split' but missing train/val_train rows.")
        if not te_rows:
            print("[GRADE] WARNING: grade_manifest split has no test rows; internal test will be skipped.", flush=True)
    else:
        # Patient-aware stratified split (fallback; use any available DR-grade-like column)
        tr_rows, va_rows, te_rows = split_rows_stratified(
            rows,
            ["label", "grade", "dr_grade", "icdr"],
            cfg.seed,
            test_size=0.1,
            val_size=0.2,
        )

    # Collect train labels for dynamic resampling (o_O-style)
    keys = list(tr_rows[0].keys()) if tr_rows else []
    label_key = get_first_existing_key(keys, ["label", "grade", "dr_grade", "icdr"])
    if label_key is None:
        raise ValueError(f"[GRADE] No label column found in manifest. Keys={sorted(keys)}")

    y_list: List[int] = []
    for r in tr_rows:
        v = r.get(label_key, "")
        y = -1 if (v == "" or v is None) else int(float(v))
        y_list.append(y)

    y_arr = np.array([y for y in y_list if y >= 0], dtype=np.int64)
    if y_arr.size > 0:
        # EyePACS 0..4 DR labels
        counts = np.bincount(y_arr, minlength=5)
        print(f"[GRADE] class counts (train) = {counts.tolist()}  num_classes=5")
    else:
        print("[GRADE] WARNING: no labeled samples found for grade training (after split).")

    has_encoder_init = (
        (cfg.init_encoder_ckpt is not None and cfg.init_encoder_ckpt.exists())
        or (cfg.roundtune_ckpt is not None and cfg.roundtune_ckpt.exists())
    )

    requested_backbone = str(cfg.backbone or "").lower().strip()
    uses_timm_pretrained = bool(requested_backbone) and (not requested_backbone.startswith("mae_vit")) and (not has_encoder_init)
    grade_aug = str(cfg.grade_aug).lower().strip()
    if grade_aug == "auto":
        grade_aug = "imagenet" if (has_encoder_init or uses_timm_pretrained) else "dr"
    if grade_aug not in ("dr", "imagenet"):
        raise ValueError(f"[GRADE] Unsupported --grade_aug '{grade_aug}'. Choose from: auto, dr, imagenet")

    GradeDS = GradeAlbDataset if grade_aug == "imagenet" else GradeDataset
    print(f"[GRADE] aug={grade_aug} (dataset={GradeDS.__name__})", flush=True)

    ds_tr = GradeDS(tr_rows, cfg.project_root, cfg.img_size, True)
    ds_va = GradeDS(va_rows, cfg.project_root, cfg.img_size, False)
    ds_te = GradeDS(te_rows, cfg.project_root, cfg.img_size, False) if te_rows else None
    ds_calib = GradeDS(calib_rows, cfg.project_root, cfg.img_size, False) if calib_rows else None

    # dl_tr will be rebuilt each epoch with dynamic WeightedRandomSampler (see below)
    dl_va = DataLoader(
        ds_va,
        batch_size=cfg.batch_size,
        shuffle=False,
        num_workers=cfg.num_workers,
        pin_memory=True,
        drop_last=False,
        persistent_workers=(cfg.num_workers > 0),
    )
    dl_te = None
    if ds_te is not None:
        dl_te = DataLoader(
            ds_te,
            batch_size=cfg.batch_size,
            shuffle=False,
            num_workers=cfg.num_workers,
            pin_memory=True,
            drop_last=False,
            persistent_workers=(cfg.num_workers > 0),
        )

    dl_calib = None
    if ds_calib is not None:
        dl_calib = DataLoader(
            ds_calib,
            batch_size=cfg.batch_size,
            shuffle=False,
            num_workers=cfg.num_workers,
            pin_memory=True,
            drop_last=False,
            persistent_workers=(cfg.num_workers > 0),
        )
    # WeightedRandomSampler for dynamic resampling in grading (o_O-style)
    try:
        from torch.utils.data import WeightedRandomSampler  # type: ignore
    except Exception:
        WeightedRandomSampler = None

    backbone_name = choose_backbone_name(cfg)
    cfg.backbone = backbone_name
    use_timm_pretrained = (not has_encoder_init) and (not backbone_name.lower().startswith("mae_vit"))
    is_transfer_run = is_grade_transfer_run(
        has_encoder_init=has_encoder_init,
        use_timm_pretrained=use_timm_pretrained,
    )
    backbone = SharedBackbone(
        backbone_name,
        pretrained=use_timm_pretrained,
        timm_pretrained_file=cfg.timm_pretrained_file,
    )
    init_info = init_backbone(backbone, cfg)
    if use_timm_pretrained and init_info.get("init") == "scratch":
        init_info = {
            **init_info,
            "init": "timm_imagenet",
            "backbone": backbone_name,
            "timm_pretrained_file": str(cfg.timm_pretrained_file) if cfg.timm_pretrained_file is not None else None,
        }

    # NOTE: o_O used an MSE regression head (1 output) trained from scratch on a CNN.
    # For transfer learning we keep MSE as default, but allow CE/Focal classification.
    grade_loss = str(cfg.grade_loss).lower().strip()
    use_regression = (grade_loss == "mse")
    if use_regression:
        num_classes_grade = 1
        print("[GRADE] mode=regression loss=mse", flush=True)
    else:
        num_classes_grade = 5
        print(f"[GRADE] mode=classification loss={grade_loss}", flush=True)

    model = GradeModel(backbone, num_classes=num_classes_grade, dropout=0.2).to(cfg.device)
    init_info = {**init_info, "finetune": apply_finetune_strategy(model, cfg)}

    is_mae = cfg.backbone.lower().startswith("mae_vit")
    if is_transfer_run:
        # Transfer learning: use AdamW + smaller backbone LR via param groups (stable for MAE/ViT/ConvNeXt).
        opt = build_optimizer(model, cfg)
        print(f"[GRADE] transfer optimizer=AdamW lr={cfg.lr:.2e} backbone_lr={cfg.lr*0.1:.2e}", flush=True)
    else:
        # From scratch: keep o_O-like SGD defaults (but still use smaller backbone LR).
        if is_mae:
            lr_grade = float(cfg.lr)
            wd_grade = float(cfg.weight_decay)
        else:
            lr_grade = 3e-3 if abs(cfg.lr - 3e-4) < 1e-10 else float(cfg.lr)
            wd_grade = 5e-4 if abs(cfg.weight_decay - 1e-4) < 1e-10 else float(cfg.weight_decay)

        backbone_params: List[torch.nn.Parameter] = []
        head_params: List[torch.nn.Parameter] = []
        for n, p in model.named_parameters():
            if not p.requires_grad:
                continue
            (backbone_params if n.startswith("backbone.") else head_params).append(p)

        opt = torch.optim.SGD(
            [
                {"params": backbone_params, "lr": lr_grade * 0.1},
                {"params": head_params, "lr": lr_grade},
            ],
            momentum=0.9,
            nesterov=True,
            weight_decay=wd_grade,
        )
    # Dynamic LR schedule: reduce LR when the target metric plateaus.
    # For grading we monitor val_qwk (fallback to val_acc if QWK is NaN).
    lr_plateau_patience = max(1, int(round(max(0, int(cfg.patience)) / 3.0)))  # e.g. patience=10 -> 3
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        opt,
        mode="max",
        factor=0.5,
        patience=lr_plateau_patience,
        threshold=float(cfg.min_delta),
        threshold_mode="rel",
        cooldown=0,
        min_lr=1e-6,
    )
    print(
        f"[GRADE] lr_scheduler=plateau monitor=val_qwk factor=0.5 patience={lr_plateau_patience} min_lr=1e-6",
        flush=True,
    )
    scaler = amp_grad_scaler(cfg.device, cfg.amp)
    writer = build_writer(cfg.run_dir, cfg.tensorboard)

    logs_dir = cfg.run_dir / "logs"
    hist_csv = logs_dir / "history.csv"
    hist_jsonl = logs_dir / "history.jsonl"
    summary_json = logs_dir / "summary.json"
    fields = ["task", "epoch", "split", "loss", "acc", "macro_f1", "qwk", "time_sec"]
    init_history_file(hist_csv, fields)
    save_json(
        summary_json,
        {
            "task": "grade",
            "init": init_info,
            "config": cfg.__dict__,
            "grade_manifest": str(grade_manifest),
        },
    )

    best_score = -1.0
    best_epoch = 0
    best_model = cfg.run_dir / "best_grade.pth"
    # Early-stop monitor (separate from best checkpoint metric).
    # Best checkpoint stays on val_qwk (research metric); early stop uses a smoother composite score.
    best_stop_score = -1.0

    patience = max(0, int(cfg.patience))
    min_delta = float(cfg.min_delta)
    bad_epochs = 0

    grade_sampler = str(cfg.grade_sampler).lower().strip()
    # Legacy single-stage sampler (kept for backward compatibility when --grade_sampler is explicitly set).
    grade_sampler_legacy = grade_sampler
    if grade_sampler_legacy == "auto":
        grade_sampler_legacy = "balanced" if is_transfer_run else "oo"
    if grade_sampler_legacy not in ("oo", "balanced", "none"):
        raise ValueError(f"[GRADE] Unsupported --grade_sampler '{grade_sampler_legacy}'. Choose from: auto, oo, balanced, none")

    # Two-stage sampler (default used when --grade_sampler=auto):
    # stage1 learns minority patterns; stage2 switches back to true prior (no WeightedRandomSampler) for calibration/DCA.
    sampler_stage1 = str(getattr(cfg, "sampler_stage1", "balanced")).lower().strip()
    sampler_stage2 = str(getattr(cfg, "sampler_stage2", "none")).lower().strip()
    if sampler_stage1 not in ("oo", "balanced"):
        raise ValueError(f"[GRADE] Unsupported --sampler_stage1 '{sampler_stage1}'. Choose from: oo, balanced")
    if sampler_stage2 not in ("none",):
        raise ValueError(f"[GRADE] Unsupported --sampler_stage2 '{sampler_stage2}'. Choose from: none")
    stage2_start_frac = float(getattr(cfg, "sampler_stage2_start_frac", 0.8))
    stage2_start_frac = float(np.clip(stage2_start_frac, 0.0, 1.0))
    stage2_start_epoch = int(np.floor(float(cfg.epochs) * stage2_start_frac))
    stage2_start_epoch = max(1, min(int(cfg.epochs), stage2_start_epoch))

    use_two_stage = (grade_sampler == "auto")
    if not use_two_stage:
        sampler_stage1 = grade_sampler_legacy
        sampler_stage2 = grade_sampler_legacy
        stage2_start_epoch = cfg.epochs + 1  # never switch

    print(
        f"[GRADE] sampler_schedule={'two_stage' if use_two_stage else 'single'} "
        f"stage1={sampler_stage1} stage2={sampler_stage2} stage2_start_epoch={stage2_start_epoch}/{cfg.epochs}",
        flush=True,
    )

    early_stop_min_epoch = max(0, int(getattr(cfg, "early_stop_min_epoch", 0)))
    if use_two_stage:
        # Only start early-stopping once stage2 begins (calibration/DCA phase).
        early_stop_min_epoch = max(int(early_stop_min_epoch), int(stage2_start_epoch))
    print(f"[GRADE] early_stop_min_epoch={early_stop_min_epoch}", flush=True)

    balanced_w: Optional[np.ndarray] = None
    if y_arr.size > 0:
        cnt = np.bincount(y_arr, minlength=5).astype(np.float32)
        cnt = np.maximum(cnt, 1.0)
        w_cls = np.sqrt(float(cnt.sum()) / cnt)
        w_cls = w_cls / float(w_cls.min())
        w_cls = np.clip(w_cls, 0.0, 10.0)
        balanced_w = w_cls.astype(np.float32)
        if (sampler_stage1 == "balanced") or (sampler_stage2 == "balanced"):
            print(f"[GRADE] sampler=balanced class_w={balanced_w.tolist()} (cap=10)", flush=True)
    elif (sampler_stage1 == "balanced") or (sampler_stage2 == "balanced"):
        print("[GRADE] WARNING: sampler=balanced requested but no labels found to compute weights.", flush=True)

    ce_weight: Optional[torch.Tensor] = None
    if (not use_regression) and y_arr.size > 0 and ((sampler_stage1 == "none") or (sampler_stage2 == "none")):
        cnt = np.bincount(y_arr, minlength=5).astype(np.float32)
        cnt = np.maximum(cnt, 1.0)
        w_cls = np.sqrt(float(cnt.sum()) / cnt)
        w_cls = w_cls / float(w_cls.mean())
        w_cls = np.clip(w_cls, 0.0, 10.0)
        ce_weight = torch.as_tensor(w_cls, dtype=torch.float32, device=cfg.device)
        print(f"[GRADE] class_weight for {grade_loss}: {w_cls.tolist()} (cap=10)", flush=True)

    train_use_ce_weight = bool(getattr(cfg, "train_use_ce_weight", True))
    val_use_ce_weight = bool(getattr(cfg, "val_use_ce_weight", False))
    if not use_regression:
        print(
            f"[GRADE] ce_weight: train_use={train_use_ce_weight} val_use={val_use_ce_weight} (only applies when sampler=none)",
            flush=True,
        )

    k_grade = 5
    qwk_w = (np.subtract.outer(np.arange(k_grade), np.arange(k_grade)) ** 2) / float((k_grade - 1) ** 2)
    def qwk_from_confusion(cm: np.ndarray) -> float:
        cm_f = cm.astype(np.float64)
        n = float(cm_f.sum())
        if n <= 0:
            return float("nan")
        hist_true = cm_f.sum(axis=1)
        if int(np.count_nonzero(hist_true)) <= 1:
            return float("nan")
        hist_pred = cm_f.sum(axis=0)
        expected = np.outer(hist_true, hist_pred) / n
        den = float((qwk_w * expected).sum())
        if den <= 0:
            return float("nan")
        num = float((qwk_w * cm_f).sum())
        return float(1.0 - (num / den))

    def build_blend_features(rows_all: List[Dict[str, str]], tta: int = 10) -> Tuple[np.ndarray, np.ndarray, List[str], List[int]]:
        """Compute per-image mean/std of regression scores with TTA; returns features, labels, patient_ids, eye_side."""
        model.eval()
        by_key: Dict[str, List[float]] = {}
        by_lab: Dict[str, int] = {}
        by_pid: Dict[str, str] = {}
        by_side: Dict[str, int] = {}

        # Use training augmentations for TTA
        ds_tta = GradeDS(rows_all, cfg.project_root, cfg.img_size, True)
        loader = DataLoader(ds_tta, batch_size=cfg.batch_size, shuffle=False, num_workers=cfg.num_workers,
                            pin_memory=True, drop_last=False, persistent_workers=(cfg.num_workers > 0))
        for _ in range(tta):
            with torch.no_grad():
                for batch in loader:
                    x = batch["image"].to(cfg.device, non_blocking=True)
                    y = batch["label"].numpy()
                    has_label = batch["has_label"].numpy()
                    paths = batch["path"]
                    scores, _ = model(x)
                    scores = scores.view(-1).detach().cpu().numpy()
                    for sc, lab, hl, p in zip(scores, y, has_label, paths):
                        if hl <= 0:
                            continue
                        by_key.setdefault(p, []).append(float(sc))
                        by_lab[p] = int(lab)
                        by_pid[p] = extract_patient_id(p)
                        by_side[p] = infer_eye_side(p)
        feats_out = []
        labels_out = []
        pids_out = []
        sides_out = []
        for p, scores in by_key.items():
            scores_np = np.array(scores, dtype=np.float32)
            mu = float(scores_np.mean())
            sigma = float(scores_np.std())
            feats_out.append((mu, sigma))
            labels_out.append(by_lab[p])
            pids_out.append(by_pid[p])
            sides_out.append(by_side[p])
        feats_arr = np.array(feats_out, dtype=np.float32)  # (N,2): mu_this, sigma_this
        labels_arr = np.array(labels_out, dtype=np.int64)
        return feats_arr, labels_arr, pids_out, sides_out

    def build_blend_dataset(feats_arr: np.ndarray, labels_arr: np.ndarray, pids: List[str], sides: List[int]) -> Tuple[np.ndarray, np.ndarray, List[str]]:
        """Construct patient-level features: [mu_this, mu_other, sigma_this, sigma_other, delta_right]."""
        feats_out = []
        labels_out = []
        groups_out = []
        by_pid_side: Dict[Tuple[str, int], List[int]] = {}
        for idx, (pid, side) in enumerate(zip(pids, sides)):
            by_pid_side.setdefault((pid, side), []).append(idx)
        for idx, (pid, side) in enumerate(zip(pids, sides)):
            mu_this, sigma_this = feats_arr[idx]
            other_side = 1 - side
            other_idxs = by_pid_side.get((pid, other_side), [])
            if other_idxs:
                j = other_idxs[0]
                mu_other, sigma_other = feats_arr[j]
            else:
                mu_other, sigma_other = 0.0, 0.0
            feats_out.append([mu_this, mu_other, sigma_this, sigma_other, float(side)])
            labels_out.append(labels_arr[idx])
            groups_out.append(pid)
        feats_flat = np.stack(feats_out, axis=0)
        feats_flat = feats_flat.reshape(feats_flat.shape[0], -1)
        return feats_flat.astype(np.float32), np.array(labels_out, dtype=np.int64), groups_out

    class BlendMLP(nn.Module):
        def __init__(self, in_dim: int = 5):
            super().__init__()
            self.net = nn.Sequential(
                nn.Linear(in_dim, 32),
                nn.ReLU(inplace=True),
                nn.Linear(32, 16),
                nn.ReLU(inplace=True),
                nn.Linear(16, 1),
            )
        def forward(self, x: torch.Tensor) -> torch.Tensor:
            return self.net(x).squeeze(1)

    def extract_rms_features(rows_all: List[Dict[str, str]], tta: int, target_dim: int = 2048) -> Tuple[np.ndarray, np.ndarray, List[str], List[int]]:
        """Extract per-image RMSPool features with TTA, pad/truncate to target_dim, return (mu,sigma), labels, patient_id, eye_side."""
        model.eval()
        by_key: Dict[str, List[np.ndarray]] = {}
        by_lab: Dict[str, int] = {}
        by_pid: Dict[str, str] = {}
        by_side: Dict[str, int] = {}
        ds_tta = GradeDS(rows_all, cfg.project_root, cfg.img_size, True)
        loader = DataLoader(ds_tta, batch_size=cfg.batch_size, shuffle=False, num_workers=cfg.num_workers,
                            pin_memory=True, drop_last=False, persistent_workers=(cfg.num_workers > 0))
        for _ in range(tta):
            with torch.no_grad():
                for batch in loader:
                    x = batch["image"].to(cfg.device, non_blocking=True)
                    y = batch["label"].numpy()
                    has_label = batch["has_label"].numpy()
                    paths = batch["path"]
                    feats, _ = model.backbone(x)  # feats list
                    fmap = feats[-1]
                    rms = torch.sqrt(torch.mean(fmap ** 2, dim=(2, 3))).detach().cpu().numpy()  # (B, C)
                    for sc, lab, hl, p in zip(rms, y, has_label, paths):
                        if hl <= 0:
                            continue
                        # pad / truncate to target_dim
                        if sc.shape[0] >= target_dim:
                            vec = sc[:target_dim]
                        else:
                            vec = np.pad(sc, (0, target_dim - sc.shape[0]), mode="constant")
                        by_key.setdefault(p, []).append(vec.astype(np.float32))
                        by_lab[p] = int(lab)
                        by_pid[p] = extract_patient_id(p)
                        by_side[p] = infer_eye_side(p)
        feats_out = []
        labels_out = []
        pids_out = []
        sides_out = []
        for p, vecs in by_key.items():
            arr = np.stack(vecs, axis=0)  # (T, D)
            mu = arr.mean(axis=0)
            sigma = arr.std(axis=0)
            feats_out.append((mu, sigma))
            labels_out.append(by_lab[p])
            pids_out.append(by_pid[p])
            sides_out.append(by_side[p])
        feats_arr = np.array(feats_out, dtype=np.float32)  # (N,2,target_dim) collapsed later
        labels_arr = np.array(labels_out, dtype=np.int64)
        return feats_arr, labels_arr, pids_out, sides_out

    for epoch in range(1, cfg.epochs + 1):
        model.train()
        t0 = time.time()
        opt.zero_grad(set_to_none=True)

        sampler_epoch = sampler_stage1 if epoch < stage2_start_epoch else sampler_stage2
        if use_two_stage and epoch == stage2_start_epoch:
            print(
                f"[GRADE] sampler switch at epoch {epoch}: {sampler_stage1}->{sampler_stage2} (stage2=true_prior)",
                flush=True,
            )
            # Stage2: reduce LR one more step for stability / smoother convergence.
            for pg in opt.param_groups:
                if "lr" in pg:
                    pg["lr"] = float(pg["lr"]) * 0.3
            lrs_now = [float(pg.get("lr", 0.0)) for pg in opt.param_groups]
            print(f"[GRADE] stage2 lr downscale x0.3 -> {lrs_now}", flush=True)

            # Stage1 often optimizes sensitivity via re-sampling and may produce over-confident scores.
            # Reset best trackers at stage2 so the final best checkpoint comes from true-prior training (DCA-valid).
            stage1_best_model = cfg.run_dir / "best_grade_stage1.pth"
            if best_model.exists():
                try:
                    best_model.replace(stage1_best_model)
                    print(f"[GRADE] archived stage1 best -> {stage1_best_model}", flush=True)
                except Exception as e:
                    print(f"[GRADE] WARNING: failed to archive stage1 best checkpoint: {e}", flush=True)
            best_score = -1.0
            best_epoch = 0
            best_stop_score = -1.0
            bad_epochs = 0
            print("[GRADE] reset best_score/best_stop_score for stage2 selection", flush=True)

        # === Build epoch-specific train loader (sampler strategy) ===
        if (sampler_epoch != "none") and (WeightedRandomSampler is not None) and (len(y_list) > 0):
            # Initial and final resampling weights as reported by o_O (EyePACS DR 0鈥?)
            w0 = np.array([1.36, 14.4, 6.64, 40.2, 49.6], dtype=np.float32)
            wf = np.array([1.0, 2.0, 2.0, 2.0, 2.0], dtype=np.float32)
            r_decay = 0.975
            alpha = float(r_decay ** max(0, epoch - 1))
            w_t = alpha * w0 + (1.0 - alpha) * wf
            w_t = np.maximum(w_t, 0.0)

            if sampler_epoch == "balanced":
                w_t = balanced_w
            elif sampler_epoch == "oo":
                pass
            else:
                w_t = None

            if w_t is None:
                dl_tr = DataLoader(
                    ds_tr,
                    batch_size=cfg.batch_size,
                    shuffle=True,
                    num_workers=cfg.num_workers,
                    pin_memory=True,
                    drop_last=True,
                    persistent_workers=(cfg.num_workers > 0),
                )
            else:
                sample_weights = []
                for y_val in y_list:
                    if 0 <= y_val < len(w_t):
                        sample_weights.append(float(w_t[y_val]))
                    else:
                        sample_weights.append(0.0)
                weights_tensor = torch.as_tensor(sample_weights, dtype=torch.double)
                sampler = WeightedRandomSampler(weights_tensor, num_samples=len(sample_weights), replacement=True)
                dl_tr = DataLoader(
                    ds_tr,
                    batch_size=cfg.batch_size,
                    sampler=sampler,
                    shuffle=False,
                    num_workers=cfg.num_workers,
                    pin_memory=True,
                    drop_last=True,
                    persistent_workers=(cfg.num_workers > 0),
                )
        else:
            if WeightedRandomSampler is None and sampler_epoch != "none":
                print("[GRADE] WeightedRandomSampler not available, falling back to shuffle=True.", flush=True)
            dl_tr = DataLoader(
                ds_tr,
                batch_size=cfg.batch_size,
                shuffle=True,
                num_workers=cfg.num_workers,
                pin_memory=True,
                drop_last=True,
                persistent_workers=(cfg.num_workers > 0),
            )

        losses = []
        loss_sum = 0.0
        steps_per_epoch = len(dl_tr)
        log_every = max(0, int(cfg.log_every_steps))
        tr_scores: List[torch.Tensor] = []
        tr_logits: List[torch.Tensor] = []
        tr_y: List[torch.Tensor] = []
        cm_run = np.zeros((k_grade, k_grade), dtype=np.int64)
        n_lab_seen = 0
        n_lab_correct = 0
        for step, batch in enumerate(dl_tr, 1):
            x = batch["image"].to(cfg.device, non_blocking=True)
            y = batch["label"].to(cfg.device, non_blocking=True)
            has_label = batch["has_label"].to(cfg.device, non_blocking=True)

            mask = has_label.bool().view(-1)
            with amp_autocast(cfg.device, cfg.amp):
                out, _ = model(x)
                if use_regression:
                    scores = out.view(-1)
                    y_float = y.float().view(-1)
                    has = has_label.float().view(-1)
                    loss_raw = F.mse_loss(scores, y_float, reduction="none") * has
                    denom = has.sum().clamp(min=1.0)
                    loss = (loss_raw.sum() / denom) / cfg.accumulate_steps
                else:
                    logits = out
                    if mask.any():
                        logits_lab = logits[mask]
                        y_lab = y.view(-1)[mask]
                        w_train = ce_weight if (sampler_epoch == "none" and train_use_ce_weight) else None
                        if grade_loss == "ce":
                            loss = F.cross_entropy(logits_lab, y_lab, weight=w_train, label_smoothing=0.1) / cfg.accumulate_steps
                        elif grade_loss == "focal":
                            loss = focal_loss_logits(logits_lab, y_lab, gamma=float(cfg.grade_focal_gamma), weight=w_train) / cfg.accumulate_steps
                        elif grade_loss == "corn":
                            loss = corn_loss(logits_lab, y_lab, num_classes=k_grade, weight=w_train) / cfg.accumulate_steps
                        elif grade_loss == "ce_qwk":
                            loss = combined_ce_qwk_loss(logits_lab, y_lab, num_classes=k_grade, qwk_weight=cfg.qwk_loss_weight, weight=w_train) / cfg.accumulate_steps
                        elif grade_loss == "ordinal":
                            loss = ordinal_regression_loss(logits_lab, y_lab, num_classes=k_grade, weight=w_train) / cfg.accumulate_steps
                        elif grade_loss == "edl":
                            anneal_epochs = int(getattr(cfg, "edl_anneal_epochs", 10) or 0)
                            if anneal_epochs > 0:
                                anneal = min(1.0, float(epoch) / float(max(1, anneal_epochs)))
                            else:
                                anneal = 1.0
                            loss = edl_digamma_loss(
                                logits_lab,
                                y_lab,
                                num_classes=k_grade,
                                evidence=str(getattr(cfg, "edl_evidence", "relu")),
                                weight=w_train,
                                kl_weight=float(getattr(cfg, "edl_kl_weight", 1.0)),
                                anneal=float(anneal),
                            ) / cfg.accumulate_steps
                        else:
                            raise ValueError(f"[GRADE] Unsupported grade_loss='{grade_loss}'")
                    else:
                        loss = torch.zeros((), dtype=torch.float32, device=cfg.device)

            if mask.any():
                if use_regression:
                    tr_scores.append(scores.detach()[mask].float().cpu())
                else:
                    tr_logits.append(out.detach()[mask].float().cpu())
                tr_y.append(y.detach().view(-1)[mask].cpu())
                y_true = y.detach().view(-1)[mask].clamp(0, k_grade - 1).cpu().numpy().astype(np.int64)
                if use_regression:
                    sc = scores.detach()[mask].float().cpu().numpy()
                    y_pred = np.zeros_like(y_true, dtype=np.int64)
                    y_pred[sc >= 0.5] = 1
                    y_pred[sc >= 1.5] = 2
                    y_pred[sc >= 2.5] = 3
                    y_pred[sc >= 3.5] = 4
                else:
                    # Prediction rule by loss family
                    if grade_loss == "corn":
                        y_pred = corn_predict(out.detach()[mask], num_classes=k_grade).cpu().numpy().astype(np.int64)
                    elif grade_loss == "edl":
                        probs_b, _, _ = edl_dirichlet_probs(out.detach()[mask].to(torch.float32), evidence=str(getattr(cfg, "edl_evidence", "relu")))
                        y_pred = probs_b.argmax(dim=1).clamp(0, k_grade - 1).cpu().numpy().astype(np.int64)
                    else:
                        y_pred = out.detach()[mask].argmax(dim=1).clamp(0, k_grade - 1).cpu().numpy().astype(np.int64)
                np.add.at(cm_run, (y_true, y_pred), 1)
                n_lab_seen += int(y_true.size)
                n_lab_correct += int((y_true == y_pred).sum())

            scaler.scale(loss).backward()
            if (step % cfg.accumulate_steps) == 0:
                if cfg.grad_clip > 0:
                    scaler.unscale_(opt)
                    nn.utils.clip_grad_norm_(model.parameters(), cfg.grad_clip)
                scaler.step(opt)
                scaler.update()
                opt.zero_grad(set_to_none=True)

            loss_item = float(loss.item()) * cfg.accumulate_steps
            losses.append(loss_item)
            loss_sum += loss_item
            if log_every > 0 and (step == 1 or step % log_every == 0 or step == steps_per_epoch):
                elapsed = time.time() - t0
                eta_sec = (elapsed / max(1, step)) * (steps_per_epoch - step)
                lr_now = float(opt.param_groups[-1].get("lr", cfg.lr)) if opt.param_groups else float(cfg.lr)
                label_frac = float(has_label.float().mean().detach().cpu().item())
                acc_run = float(n_lab_correct / max(1, n_lab_seen)) if n_lab_seen > 0 else 0.0
                qwk_run = qwk_from_confusion(cm_run) if n_lab_seen > 0 else float("nan")
                print(
                    f"[GRADE][Epoch {epoch}/{cfg.epochs}][{step}/{steps_per_epoch}] "
                    f"loss={loss_item:.4f} avg={loss_sum / step:.4f} acc={acc_run:.4f} qwk={qwk_run:.4f} "
                    f"lab={label_frac:.2f} lr={lr_now:.2e} eta={eta_sec/60:.1f}m",
                    flush=True,
                )

        # If the last few batches didn't trigger an optimizer step (steps_per_epoch not divisible by accumulate_steps),
        # apply one final step so those gradients aren't dropped.
        if (cfg.accumulate_steps > 1) and ((steps_per_epoch % cfg.accumulate_steps) != 0):
            if cfg.grad_clip > 0:
                scaler.unscale_(opt)
                nn.utils.clip_grad_norm_(model.parameters(), cfg.grad_clip)
            scaler.step(opt)
            scaler.update()
            opt.zero_grad(set_to_none=True)

        train_loss = float(np.mean(losses)) if losses else 0.0
        train_acc, train_f1, train_qwk = 0.0, np.nan, np.nan
        if tr_y:
            if use_regression:
                mets_tr = grade_regression_metrics(torch.cat(tr_scores, dim=0), torch.cat(tr_y, dim=0).long())
            else:
                if grade_loss == "corn":
                    mets_tr = cls_metrics_corn(torch.cat(tr_logits, dim=0), torch.cat(tr_y, dim=0).long(), num_classes=k_grade)
                elif grade_loss == "edl":
                    logits_cat = torch.cat(tr_logits, dim=0).to(torch.float32)
                    probs_cat, _, _ = edl_dirichlet_probs(logits_cat, evidence=str(getattr(cfg, "edl_evidence", "relu")))
                    mets_tr = cls_metrics_from_probs(probs_cat, torch.cat(tr_y, dim=0).long())
                else:
                    mets_tr = cls_metrics(torch.cat(tr_logits, dim=0), torch.cat(tr_y, dim=0).long())
            train_acc = float(mets_tr.get("acc", 0.0))
            train_f1 = float(mets_tr.get("macro_f1", np.nan))
            train_qwk = float(mets_tr.get("qwk", np.nan))

        # === Validation (regression 鈫?ordinal levels 鈫?QWK/Acc) ===
        model.eval()
        all_scores, all_logits, all_y, val_losses = [], [], [], []
        val_loss_unw_sum = 0.0
        val_loss_w_sum = 0.0
        val_loss_n = 0
        with torch.no_grad():
            for batch in dl_va:
                x = batch["image"].to(cfg.device, non_blocking=True)
                y = batch["label"].to(cfg.device, non_blocking=True)
                has_label = batch["has_label"].to(cfg.device, non_blocking=True).bool()
                out, _ = model(x)
                mask = has_label.view(-1)
                if not mask.any():
                    continue
                if use_regression:
                    scores = out.view(-1)
                    has = has_label.float().view(-1)
                    loss_raw = F.mse_loss(scores, y.float().view(-1), reduction="none") * has
                    all_scores.append(scores[has_label].detach().cpu())
                    all_y.append(y[has_label].detach().cpu())
                    denom = has_label.float().sum().clamp(min=1.0)
                    val_losses.append(float((loss_raw.sum() / denom).item()))
                else:
                    logits = out
                    logits_lab = logits[mask]
                    y_lab = y.view(-1)[mask]
                    if grade_loss == "ce":
                        loss_unw = F.cross_entropy(logits_lab, y_lab, weight=None, reduction="mean", label_smoothing=0.1)
                        if val_use_ce_weight and (ce_weight is not None):
                            loss_w = F.cross_entropy(logits_lab, y_lab, weight=ce_weight, reduction="mean", label_smoothing=0.1)
                        else:
                            loss_w = loss_unw
                    elif grade_loss == "focal":
                        loss_unw = focal_loss_logits(logits_lab, y_lab, gamma=float(cfg.grade_focal_gamma), weight=None, reduction="mean")
                        if val_use_ce_weight and (ce_weight is not None):
                            loss_w = focal_loss_logits(logits_lab, y_lab, gamma=float(cfg.grade_focal_gamma), weight=ce_weight, reduction="mean")
                        else:
                            loss_w = loss_unw
                    elif grade_loss == "corn":
                        loss_unw = corn_loss(logits_lab, y_lab, num_classes=k_grade, weight=None)
                        if val_use_ce_weight and (ce_weight is not None):
                            loss_w = corn_loss(logits_lab, y_lab, num_classes=k_grade, weight=ce_weight)
                        else:
                            loss_w = loss_unw
                    elif grade_loss == "ce_qwk":
                        loss_unw = combined_ce_qwk_loss(logits_lab, y_lab, num_classes=k_grade, qwk_weight=cfg.qwk_loss_weight, weight=None)
                        if val_use_ce_weight and (ce_weight is not None):
                            loss_w = combined_ce_qwk_loss(logits_lab, y_lab, num_classes=k_grade, qwk_weight=cfg.qwk_loss_weight, weight=ce_weight)
                        else:
                            loss_w = loss_unw
                    elif grade_loss == "ordinal":
                        loss_unw = ordinal_regression_loss(logits_lab, y_lab, num_classes=k_grade, weight=None)
                        if val_use_ce_weight and (ce_weight is not None):
                            loss_w = ordinal_regression_loss(logits_lab, y_lab, num_classes=k_grade, weight=ce_weight)
                        else:
                            loss_w = loss_unw
                    elif grade_loss == "edl":
                        anneal_epochs = int(getattr(cfg, "edl_anneal_epochs", 10) or 0)
                        if anneal_epochs > 0:
                            anneal = min(1.0, float(epoch) / float(max(1, anneal_epochs)))
                        else:
                            anneal = 1.0
                        loss_unw = edl_digamma_loss(
                            logits_lab,
                            y_lab,
                            num_classes=k_grade,
                            evidence=str(getattr(cfg, "edl_evidence", "relu")),
                            weight=None,
                            kl_weight=float(getattr(cfg, "edl_kl_weight", 1.0)),
                            anneal=float(anneal),
                        )
                        if val_use_ce_weight and (ce_weight is not None):
                            loss_w = edl_digamma_loss(
                                logits_lab,
                                y_lab,
                                num_classes=k_grade,
                                evidence=str(getattr(cfg, "edl_evidence", "relu")),
                                weight=ce_weight,
                                kl_weight=float(getattr(cfg, "edl_kl_weight", 1.0)),
                                anneal=float(anneal),
                            )
                        else:
                            loss_w = loss_unw
                    else:
                        raise ValueError(f"[GRADE] Unsupported grade_loss='{grade_loss}'")
                    bs = int(mask.sum().item())
                    val_loss_unw_sum += float(loss_unw.item()) * bs
                    val_loss_w_sum += float(loss_w.item()) * bs
                    val_loss_n += bs
                    all_logits.append(logits_lab.detach().cpu())
                    all_y.append(y_lab.detach().cpu())

        mets: Dict[str, float] = {}
        if all_y:
            y_cat = torch.cat(all_y, dim=0).long()
            if use_regression:
                scores_cat = torch.cat(all_scores, dim=0)
                mets = grade_regression_metrics(scores_cat, y_cat)
            else:
                logits_cat = torch.cat(all_logits, dim=0)
                if grade_loss == "corn":
                    mets = cls_metrics_corn(logits_cat, y_cat, num_classes=k_grade)
                elif grade_loss == "edl":
                    probs_cat, _, _ = edl_dirichlet_probs(logits_cat.to(torch.float32), evidence=str(getattr(cfg, "edl_evidence", "relu")))
                    mets = cls_metrics_from_probs(probs_cat, y_cat)
                else:
                    mets = cls_metrics(logits_cat, y_cat)
        val_acc = float(mets.get("acc", 0.0)) if mets else 0.0
        val_f1 = float(mets.get("macro_f1", np.nan)) if mets else np.nan
        val_qwk = float(mets.get("qwk", np.nan)) if mets else np.nan
        if use_regression:
            val_loss_unw = float(np.mean(val_losses)) if val_losses else 0.0
            val_loss_w = val_loss_unw
        else:
            val_loss_unw = float(val_loss_unw_sum / max(1, val_loss_n)) if val_loss_n > 0 else 0.0
            val_loss_w = float(val_loss_w_sum / max(1, val_loss_n)) if val_loss_n > 0 else val_loss_unw

        dt = time.time() - t0
        print(
            f"[GRADE][Epoch {epoch}/{cfg.epochs}] train_loss={train_loss:.4f} "
            f"val_loss_unw={val_loss_unw:.4f} val_loss_w={val_loss_w:.4f} "
            f"val_acc={val_acc:.4f} val_qwk={val_qwk:.4f} time={dt:.1f}s",
            flush=True,
        )

        append_history_row(hist_csv, {"task": "grade", "epoch": epoch, "split": "train", "loss": train_loss, "acc": train_acc, "macro_f1": train_f1, "qwk": train_qwk, "time_sec": dt}, fields)
        append_history_row(hist_csv, {"task": "grade", "epoch": epoch, "split": "val", "loss": val_loss_unw, "acc": val_acc, "macro_f1": val_f1, "qwk": val_qwk, "time_sec": ""}, fields)
        append_jsonl(hist_jsonl, {"task": "grade", "epoch": epoch, "train_loss": train_loss, "train_acc": train_acc, "train_macro_f1": train_f1, "train_qwk": train_qwk, "val_loss": val_loss_unw, "val_loss_w": val_loss_w, "val_acc": val_acc, "val_macro_f1": val_f1, "val_qwk": val_qwk, "time_sec": dt})

        if writer is not None:
            writer.add_scalar("grade/train_loss", train_loss, epoch)
            writer.add_scalar("grade/train_acc", train_acc, epoch)
            if not np.isnan(train_f1):
                writer.add_scalar("grade/train_macro_f1", train_f1, epoch)
            if not np.isnan(train_qwk):
                writer.add_scalar("grade/train_qwk", train_qwk, epoch)
            writer.add_scalar("grade/val_loss_unw", val_loss_unw, epoch)
            writer.add_scalar("grade/val_loss_w", val_loss_w, epoch)
            writer.add_scalar("grade/val_acc", val_acc, epoch)
            if not np.isnan(val_f1):
                writer.add_scalar("grade/val_macro_f1", val_f1, epoch)
            if not np.isnan(val_qwk):
                writer.add_scalar("grade/val_qwk", val_qwk, epoch)

        # Best checkpoint metric: QWK (fallback to acc if QWK is undefined).
        best_metric = float(val_qwk) if not np.isnan(val_qwk) else val_acc
        if best_metric > (best_score + min_delta):
            best_score = best_metric
            best_epoch = epoch
            torch.save(model.state_dict(), best_model)
            save_json(
                summary_json,
                {
                    "task": "grade",
                    "best_score": best_score,
                    "best_epoch": best_epoch,
                    "last_epoch": epoch,
                    "config": cfg.__dict__,
                    "init": init_info,
                    "grade_manifest": str(grade_manifest),
                },
            )
            print(f"[GRADE] [OK] Saved best: val_score={best_score:.4f}")

        # Early-stop metric: prioritize QWK (fallback to acc if QWK is undefined).
        stop_score = float(val_qwk) if not np.isnan(val_qwk) else val_acc
        if epoch < early_stop_min_epoch:
            # Guard: do not update bad_epochs / stop before early_stop_min_epoch.
            if stop_score > (best_stop_score + min_delta):
                best_stop_score = stop_score
        else:
            if stop_score > (best_stop_score + min_delta):
                best_stop_score = stop_score
                bad_epochs = 0
            else:
                bad_epochs += 1
                if patience > 0 and bad_epochs >= patience:
                    print(
                        f"[GRADE] Early stopping at epoch {epoch} (no improvement for {bad_epochs} epochs; "
                        f"stop_score={stop_score:.4f} best_stop_score={best_stop_score:.4f}).",
                        flush=True,
                    )
                    break
        lr_before = [float(pg.get("lr", 0.0)) for pg in opt.param_groups]
        sched_score = float(val_qwk) if not np.isnan(val_qwk) else val_acc
        scheduler.step(sched_score)
        lr_after = [float(pg.get("lr", 0.0)) for pg in opt.param_groups]
        if len(lr_before) == len(lr_after) and any(abs(a - b) > 1e-12 for a, b in zip(lr_before, lr_after)):
            msg = f"[GRADE] LR adjusted: head {lr_before[-1]:.2e}->{lr_after[-1]:.2e}"
            if len(lr_before) >= 2:
                msg = f"[GRADE] LR adjusted: backbone {lr_before[0]:.2e}->{lr_after[0]:.2e} head {lr_before[-1]:.2e}->{lr_after[-1]:.2e}"
            print(msg, flush=True)
    if writer is not None:
        writer.close()

    # Load best checkpoint for calibration / final evaluation (keeps best_ckpt on val_qwk unchanged).
    if best_model.exists():
        try:
            sd = torch.load(best_model, map_location=cfg.device)
            model.load_state_dict(sd, strict=True)
            model.eval()
            print(f"[GRADE] loaded best checkpoint: {best_model}", flush=True)
        except Exception as e:
            print(f"[GRADE] WARNING: failed to load best checkpoint '{best_model}': {e}", flush=True)

    # Temperature scaling on calib split (if provided in manifest).
    temperature_json = logs_dir / "temperature.json"
    if (dl_calib is not None) and (not use_regression) and (str(getattr(cfg, "grade_loss", "")).lower().strip() != "edl"):
        print(f"[GRADE][Calib] fitting temperature on calib split (n={len(calib_rows)})", flush=True)
        calib_logits: List[torch.Tensor] = []
        calib_y: List[torch.Tensor] = []
        with torch.no_grad():
            for batch in dl_calib:
                x = batch["image"].to(cfg.device, non_blocking=True)
                y = batch["label"].to(cfg.device, non_blocking=True)
                has_label = batch["has_label"].to(cfg.device, non_blocking=True).bool().view(-1)
                out, _ = model(x)
                if not has_label.any():
                    continue
                calib_logits.append(out.detach()[has_label].to(torch.float32).cpu())
                calib_y.append(y.view(-1)[has_label].detach().cpu())

        if calib_y:
            logits_all = torch.cat(calib_logits, dim=0).to(cfg.device)
            y_all = torch.cat(calib_y, dim=0).long().to(cfg.device)

            with torch.no_grad():
                nll_before = float(F.cross_entropy(logits_all, y_all, reduction="mean").item())

            log_T = torch.zeros(1, device=cfg.device, requires_grad=True)
            opt_T = torch.optim.LBFGS([log_T], lr=0.1, max_iter=50)

            def _closure() -> torch.Tensor:
                opt_T.zero_grad(set_to_none=True)
                T = torch.exp(log_T).clamp(min=1e-3, max=100.0)
                loss = F.cross_entropy(logits_all / T, y_all, reduction="mean")
                loss.backward()
                return loss

            opt_T.step(_closure)
            T = float(torch.exp(log_T).clamp(min=1e-3, max=100.0).detach().cpu().item())

            with torch.no_grad():
                nll_after = float(F.cross_entropy(logits_all / T, y_all, reduction="mean").item())

            save_json(temperature_json, {"T": T, "n_calib": int(y_all.numel()), "nll_before": nll_before, "nll_after": nll_after})
            print(f"[GRADE][Calib] saved temperature: T={T:.4f} nll={nll_before:.4f}->{nll_after:.4f} -> {temperature_json}", flush=True)
        else:
            print("[GRADE][Calib] WARNING: no labeled samples found in calib split; skipped.", flush=True)
    else:
        if dl_calib is None:
            print("[GRADE][Calib] skipped (no calib split in grade_manifest).", flush=True)
        elif use_regression:
            print("[GRADE][Calib] skipped (grade_loss=mse).", flush=True)
        else:
            print("[GRADE][Calib] skipped (grade_loss=edl: temperature scaling is not applicable).", flush=True)

    blend_metrics = {}
    te_acc, te_f1, te_qwk = 0.0, np.nan, np.nan
    all_scores, all_logits, all_y = [], [], []
    if dl_te is not None:
        model.eval()
        with torch.no_grad():
            for batch in dl_te:
                x = batch["image"].to(cfg.device, non_blocking=True)
                y = batch["label"].to(cfg.device, non_blocking=True)
                has_label = batch["has_label"].to(cfg.device, non_blocking=True).bool()
                out, _ = model(x)
                mask = has_label.view(-1)
                if not mask.any():
                    continue
                if use_regression:
                    scores = out.view(-1)
                    all_scores.append(scores[mask].detach().cpu())
                else:
                    all_logits.append(out[mask].detach().cpu())
                all_y.append(y.view(-1)[mask].detach().cpu())
    else:
        print("[GRADE] Internal test skipped (no test split).", flush=True)

    if all_y:
        y_cat = torch.cat(all_y, dim=0).long()
        if use_regression:
            scores_cat = torch.cat(all_scores, dim=0)
            mets = grade_regression_metrics(scores_cat, y_cat)
        else:
            logits_cat = torch.cat(all_logits, dim=0)
            if cfg.grade_loss == "corn":
                mets = cls_metrics_corn(logits_cat, y_cat, num_classes=k_grade)
            elif cfg.grade_loss == "edl":
                probs_cat, _, _ = edl_dirichlet_probs(logits_cat.to(torch.float32), evidence=str(getattr(cfg, "edl_evidence", "relu")))
                mets = cls_metrics_from_probs(probs_cat, y_cat)
            else:
                mets = cls_metrics(logits_cat, y_cat)
        te_acc = float(mets.get("acc", 0.0))
        te_f1 = float(mets.get("macro_f1", np.nan))
        te_qwk = float(mets.get("qwk", np.nan))

    if dl_te is not None and all_y:
        if not np.isnan(te_f1):
            print(f"[GRADE][Internal Test] acc={te_acc:.4f} macro_f1={te_f1:.4f} qwk={te_qwk:.4f}")
        else:
            print(f"[GRADE][Internal Test] acc={te_acc:.4f} qwk={te_qwk:.4f}")
    elif dl_te is not None:
        print("[GRADE] WARNING: internal test split has no labeled samples; skipped.", flush=True)

    # Optional patient-level blend (o_O style)
    if cfg.grade_blend:
        if not use_regression:
            print("[GRADE][Blend] skipped (grade_loss != mse only supports regression blend).")
        else:
            try:
                rows_all = tr_rows + va_rows + te_rows
                feats_arr, labels_arr, pids_all, sides_all = extract_rms_features(rows_all, tta=max(1, int(cfg.grade_blend_tta)), target_dim=2048)
                X, y_blend, groups = build_blend_dataset(feats_arr, labels_arr, pids_all, sides_all)

                # Split by patient
                if GroupShuffleSplit is not None:
                    gss = GroupShuffleSplit(n_splits=1, test_size=0.2, random_state=cfg.seed)
                    train_idx, val_idx = next(gss.split(X, y_blend, groups))
                else:
                    rng = np.random.RandomState(cfg.seed)
                    perm = rng.permutation(len(X))
                    n_val = max(1, int(round(len(X) * 0.2)))
                    val_idx = perm[:n_val]
                    train_idx = perm[n_val:]

                X_tr = torch.tensor(X[train_idx], dtype=torch.float32, device=cfg.device)
                y_tr_blend = torch.tensor(y_blend[train_idx], dtype=torch.float32, device=cfg.device)
                X_va = torch.tensor(X[val_idx], dtype=torch.float32, device=cfg.device)
                y_va_blend = torch.tensor(y_blend[val_idx], dtype=torch.float32, device=cfg.device)

                blend = BlendMLP(in_dim=X.shape[1]).to(cfg.device)
                opt_blend = torch.optim.Adam(blend.parameters(), lr=1e-3, weight_decay=1e-4)

                B = 128
                steps_per_epoch = max(1, int(np.ceil(len(X_tr) / B)))
                prev_idx = None
                for ep in range(1, int(cfg.grade_blend_epochs) + 1):
                    blend.train()
                    for _ in range(steps_per_epoch):
                        mode = np.random.rand()
                        if mode < 0.2:
                            # balanced
                            idxs = []
                            uniq = np.unique(y_blend[train_idx])
                            per_class = max(1, B // max(1, len(uniq)))
                            for c in uniq:
                                pool = np.where(y_blend[train_idx] == c)[0]
                                choice = np.random.choice(pool, size=per_class, replace=True)
                                idxs.extend(choice.tolist())
                            idxs = np.array(idxs[:B])
                        elif mode < 0.7:
                            # uniform
                            pool = np.arange(len(train_idx))
                            idxs = np.random.choice(pool, size=min(B, len(pool)), replace=True)
                        else:
                            if prev_idx is None:
                                pool = np.arange(len(train_idx))
                                idxs = np.random.choice(pool, size=min(B, len(pool)), replace=True)
                            else:
                                idxs = prev_idx
                        prev_idx = idxs
                        xb = torch.tensor(X_tr.cpu().numpy()[idxs], dtype=torch.float32, device=cfg.device)
                        yb = torch.tensor(y_blend[train_idx][idxs], dtype=torch.float32, device=cfg.device)
                        opt_blend.zero_grad()
                        pred = blend(xb)
                        loss = F.mse_loss(pred, yb)
                        loss.backward()
                        opt_blend.step()

                blend.eval()
                with torch.no_grad():
                    pred_va = blend(X_va)
                blend_mets = grade_regression_metrics(pred_va.detach().cpu(), y_va_blend.detach().cpu().long())
                blend_metrics = {f"Blend_{k}": float(v) for k, v in blend_mets.items()}
                print(f"[GRADE][Blend] val metrics: {blend_metrics}")
            except Exception as e:
                print(f"[GRADE][Blend] skipped due to error: {e}")

    extra = {"TestAcc": te_acc}
    if not np.isnan(te_f1):
        extra["TestMacroF1"] = float(te_f1)
    if not np.isnan(te_qwk):
        extra["TestQWK"] = float(te_qwk)
    for k, v in blend_metrics.items():
        extra[k] = v

    finalize_table(
        run_dir=cfg.run_dir,
        task_name="DR grading (image-only)",
        dataset_name="EyePACS (labeled train; internal split)",
        cfg=cfg.__dict__,
        best_epoch=best_epoch,
        best_metric="ValScore",
        best_value=best_score,
        extra=extra,
    )
    save_json(
        summary_json,
        {
            "task": "grade",
            "init": init_info,
            "config": cfg.__dict__,
            "grade_manifest": str(grade_manifest),
            "best_epoch": best_epoch,
            "best_val_score": best_score,
            "internal_test": {"acc": te_acc, "macro_f1": te_f1, "qwk": te_qwk},
            "blend": blend_metrics,
        },
    )
    print(f"[GRADE] Done. Outputs in: {cfg.run_dir}")


# -----------------------------
# CLI
# -----------------------------
def _parse_int_list(s: str) -> List[int]:
    s = (s or "").strip()
    if not s:
        return []
    out: List[int] = []
    for part in s.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        out.append(int(part))
    return out

def _parse_float_list(s: str) -> List[float]:
    s = (s or "").strip()
    if not s:
        return []
    out: List[float] = []
    for part in s.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        out.append(float(part))
    return out

# -----------------------------
# CLI
# -----------------------------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()

    p.add_argument("--task", type=str, choices=["qual", "grade"], required=True)
    p.add_argument("--project_root", type=str, required=True)
    p.add_argument("--output_root", type=str, default="checkpoints/runs",
                   help="Output base dir, absolute or relative to project_root.")
    p.add_argument("--backbone", type=str, default="",
                   help="timm model name; empty means infer from RoundTune or use convnext_tiny.")
    p.add_argument("--roundtune_ckpt", type=str, default="", help="RoundTune encoder checkpoint.")
    p.add_argument("--init_encoder_ckpt", type=str, default="", help="Encoder checkpoint, e.g. best_encoder.pth.")
    p.add_argument("--init_quality_ckpt", type=str, default="", help="Reserved for compatibility with internal runs.")
    p.add_argument("--timm_pretrained_file", type=str, default="", help="Optional local timm pretrained weight file.")
    p.add_argument("--qual_manifest", type=str, default="")
    p.add_argument("--grade_manifest", type=str, default="")

    p.add_argument("--img_size", type=int, default=512)
    p.add_argument("--batch_size", type=int, default=8)
    p.add_argument("--epochs", type=int, default=15)
    p.add_argument("--lr", type=float, default=1e-3)
    p.add_argument("--weight_decay", type=float, default=1e-4)
    p.add_argument("--dropout", type=float, default=0.2)
    p.add_argument("--num_workers", type=int, default=4)
    p.add_argument("--amp", action="store_true")
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--device", type=str, default="cuda" if torch.cuda.is_available() else "cpu")
    p.add_argument("--tensorboard", action="store_true")
    p.add_argument("--grad_clip", type=float, default=1.0)
    p.add_argument("--accumulate_steps", type=int, default=1)
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--log_every_steps", type=int, default=1)
    p.add_argument("--patience", type=int, default=0, help="Early stopping patience; 0 disables.")
    p.add_argument("--min_delta", type=float, default=0.0)
    p.add_argument("--early_stop_min_epoch", type=int, default=0)

    p.add_argument("--tune_mode", "--tune", dest="tune_mode", type=str, default="full",
                   choices=["full", "freeze", "ln", "last_n", "lora"])
    p.add_argument("--unfreeze_last_n", type=int, default=0)
    p.add_argument("--lora_r", type=int, default=0)
    p.add_argument("--lora_alpha", type=float, default=16.0)
    p.add_argument("--lora_dropout", type=float, default=0.0)
    p.add_argument("--lora_targets", type=str, default="qkv,proj")
    p.add_argument("--lora_lr_mult", type=float, default=0.1)
    p.add_argument("--lora_wd_mult", type=float, default=1.0)
    p.add_argument("--head_lr_mult", type=float, default=1.0)

    p.add_argument("--num_classes_quality", type=int, default=3)
    p.add_argument("--qual_sampler", type=str, default="auto", choices=["auto", "weighted", "none"])
    p.add_argument("--qual_gradable_index", type=int, default=1, choices=[0, 1])
    p.add_argument("--qual_collapse_to_binary", action="store_true")
    p.add_argument("--qual_no_weighted_sampler", action="store_true")
    p.add_argument("--qual_no_class_weights", action="store_true")

    p.add_argument("--grade_loss", type=str, default="mse",
                   choices=["mse", "focal", "ce", "corn", "ce_qwk", "ordinal", "edl"])
    p.add_argument("--grade_focal_gamma", type=float, default=2.0)
    p.add_argument("--qwk_loss_weight", type=float, default=0.3)
    p.add_argument("--edl_evidence", type=str, default="relu", choices=["relu", "softplus"])
    p.add_argument("--edl_kl_weight", type=float, default=1.0)
    p.add_argument("--edl_anneal_epochs", type=int, default=10)
    p.add_argument("--grade_blend", action="store_true")
    p.add_argument("--grade_blend_tta", type=int, default=10)
    p.add_argument("--grade_blend_epochs", type=int, default=20)
    p.add_argument("--grade_aug", type=str, default="auto", choices=["auto", "dr", "imagenet"])
    p.add_argument("--grade_sampler", type=str, default="auto", choices=["auto", "oo", "balanced", "none"])
    p.add_argument("--sampler_stage1", type=str, default="balanced", choices=["balanced", "oo"])
    p.add_argument("--sampler_stage2", type=str, default="none", choices=["none"])
    p.add_argument("--sampler_stage2_start_frac", type=float, default=0.8)
    p.add_argument("--train_use_ce_weight", type=int, default=0, choices=[0, 1])
    p.add_argument("--val_use_ce_weight", type=int, default=0, choices=[0, 1])

    return p.parse_args()


def main() -> None:
    args = parse_args()
    seed_everything(args.seed)

    project_root = Path(args.project_root).resolve()
    out_root = Path(resolve_path(project_root, args.output_root)).resolve()
    run_dir = out_root / args.task
    ensure_dir(run_dir)

    roundtune_arg_provided = bool(args.roundtune_ckpt)
    roundtune_ckpt = Path(resolve_path(project_root, args.roundtune_ckpt)) if args.roundtune_ckpt else None
    if roundtune_ckpt is None:
        default_ckpt = project_root / "weights" / "roundtune" / "roundtune_encoder.pth"
        roundtune_ckpt = default_ckpt if default_ckpt.exists() else None
    init_encoder_ckpt = Path(resolve_path(project_root, args.init_encoder_ckpt)) if args.init_encoder_ckpt else None
    init_quality_ckpt = Path(resolve_path(project_root, args.init_quality_ckpt)) if args.init_quality_ckpt else None
    timm_pretrained_file = Path(resolve_path(project_root, args.timm_pretrained_file)) if args.timm_pretrained_file else None

    backbone = args.backbone.strip()
    if not backbone:
        if roundtune_ckpt is not None and roundtune_ckpt.exists():
            _, inferred = load_roundtune_encoder_weights(roundtune_ckpt)
            backbone = inferred or "convnext_tiny"
            print(f"[INFO] backbone inferred from RoundTune: {backbone}")
        else:
            backbone = "convnext_tiny"
            print(f"[INFO] backbone defaulted to: {backbone}")
    elif (not roundtune_arg_provided) and (not backbone.lower().startswith("mae_vit")):
        roundtune_ckpt = None

    qual_sampler = str(getattr(args, "qual_sampler", "auto") or "auto").lower().strip()
    qual_use_weighted_sampler = not bool(getattr(args, "qual_no_weighted_sampler", False))
    if qual_sampler in ("weighted", "none"):
        qual_use_weighted_sampler = qual_sampler == "weighted"

    cfg = TrainCfg(
        task=args.task,
        project_root=project_root,
        run_dir=run_dir,
        backbone=backbone,
        roundtune_ckpt=roundtune_ckpt,
        init_encoder_ckpt=init_encoder_ckpt,
        init_quality_ckpt=init_quality_ckpt,
        timm_pretrained_file=timm_pretrained_file,
        img_size=args.img_size,
        batch_size=args.batch_size,
        epochs=args.epochs,
        lr=args.lr,
        weight_decay=args.weight_decay,
        dropout=float(args.dropout),
        num_workers=args.num_workers,
        amp=bool(args.amp),
        seed=args.seed,
        device=args.device,
        tensorboard=bool(args.tensorboard),
        grad_clip=args.grad_clip,
        accumulate_steps=max(1, args.accumulate_steps),
        limit=max(0, int(args.limit)),
        log_every_steps=max(0, int(args.log_every_steps)),
        tune_mode=str(args.tune_mode),
        unfreeze_last_n=int(args.unfreeze_last_n),
        lora_r=int(args.lora_r),
        lora_alpha=float(args.lora_alpha),
        lora_dropout=float(args.lora_dropout),
        lora_targets=str(args.lora_targets),
        lora_lr_mult=float(args.lora_lr_mult),
        lora_wd_mult=float(args.lora_wd_mult),
        head_lr_mult=float(args.head_lr_mult),
        num_classes_quality=args.num_classes_quality,
        qual_use_weighted_sampler=bool(qual_use_weighted_sampler),
        qual_use_class_weights=not bool(args.qual_no_class_weights),
        qual_gradable_index=int(args.qual_gradable_index),
        qual_collapse_to_binary=bool(args.qual_collapse_to_binary),
        grade_loss=args.grade_loss,
        grade_focal_gamma=args.grade_focal_gamma,
        qwk_loss_weight=float(args.qwk_loss_weight),
        edl_evidence=str(args.edl_evidence),
        edl_kl_weight=float(args.edl_kl_weight),
        edl_anneal_epochs=int(args.edl_anneal_epochs),
        grade_blend=bool(args.grade_blend),
        grade_blend_tta=int(args.grade_blend_tta),
        grade_blend_epochs=int(args.grade_blend_epochs),
        grade_aug=str(args.grade_aug),
        grade_sampler=str(args.grade_sampler),
        sampler_stage1=str(args.sampler_stage1),
        sampler_stage2=str(args.sampler_stage2),
        sampler_stage2_start_frac=float(args.sampler_stage2_start_frac),
        train_use_ce_weight=bool(int(args.train_use_ce_weight)),
        val_use_ce_weight=bool(int(args.val_use_ce_weight)),
        early_stop_min_epoch=max(0, int(args.early_stop_min_epoch)),
        patience=max(0, int(args.patience)),
        min_delta=float(args.min_delta),
    )
    torch.backends.cudnn.benchmark = True

    if cfg.task == "qual":
        if not args.qual_manifest:
            raise ValueError("--qual_manifest is required for --task qual")
        train_qual(cfg, Path(resolve_path(project_root, args.qual_manifest)))
    elif cfg.task == "grade":
        if not args.grade_manifest:
            raise ValueError("--grade_manifest is required for --task grade")
        train_grade(cfg, Path(resolve_path(project_root, args.grade_manifest)))
    else:
        raise ValueError(cfg.task)


if __name__ == "__main__":
    main()


if __name__ == "__main__":
    main()



